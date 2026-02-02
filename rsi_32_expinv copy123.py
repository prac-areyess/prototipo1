import pyautogui
import subprocess
import time
import os
import sys
import logging
import traceback
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from pywinauto import Application, Desktop

# Deshabilitar fail-safe de pyautogui para evitar errores cuando el mouse se mueve a las esquinas
pyautogui.FAILSAFE = False

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('proceso_log32.txt'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Rutas
if getattr(sys, 'frozen', False):
    SCRIPT_DIR = Path(sys.executable).parent
else:
    SCRIPT_DIR = Path(__file__).parent

SHORTCUT_PATH = SCRIPT_DIR / "Actualiza RSIRAT.lnk"
IMAGES_DIR = SCRIPT_DIR

logger.info(f"Directorio de trabajo: {SCRIPT_DIR}")
logger.info(f"Arquitectura Python: {'64-bit' if sys.maxsize > 2**32 else '32-bit'}")




class RSIRATAutomation32:
    """Automatización de RSIRAT optimizada para 32-bit desde Python 64-bit"""
    
    def __init__(self, confidence_threshold=0.40):
        self.password = None
        self.dependencia = None
        self.expediente = None
        self.dep_type = None  # "21" o "23" - Indica el BUCLE INICIAL a usar (no el tipo)
        self.primer_expediente_idx = 0  # Índice del primer expediente válido encontrado
        # Indica si el expediente anterior quedó completamente procesado
        # (RC extraída o MONTO MAYOR registrado). Usada para decidir
        # si se puede ejecutar `click_cambio_expediente()` antes del
        # siguiente expediente.
        self.last_exp_completed = False
        
        # Coordenadas para desplazamiento del menú
        self.trabar_embargo_coords = None  # Coordenadas de "Trabar Embargo"
        self.proceso_embargo_coords = None  # Coordenadas de "Proceso de Embargo"
    
    def validate_expediente_row(self, expedientes, row_idx):
        """
        Valida que un expediente específico tenga todos los datos necesarios.
        
        Verificaciones:
        1. DEPENDENCIA: No vacío
        2. TIPO DE MEDIDA: IEI o DSE
        3. Si IEI: INTERVENTOR y PLAZO no vacíos
        4. Si DSE: MONTO no vacío
        
        Args:
            expedientes: DataFrame del Excel
            row_idx: Índice de la fila (0-based)
        
        Retorna:
            - (True, "") si el expediente es válido
            - (False, "mensaje de error") si falta algo
        """
        try:
            fila = expedientes.iloc[row_idx]
            errores = []
            
            # ============================================================
            # 1. VALIDAR DEPENDENCIA
            # ============================================================
            if "DEPENDENCIA" not in expedientes.columns:
                errores.append("FALTA COLUMNA 'DEPENDENCIA'")
            else:
                dependencia = str(fila.get("DEPENDENCIA", "")).strip()
                if not dependencia or dependencia.upper() == "NAN":
                    errores.append("FALTA DEPENDENCIA")
            
            # ============================================================
            # 2. VALIDAR TIPO DE MEDIDA
            # ============================================================
            if "TIPO DE MEDIDA" not in expedientes.columns:
                errores.append("FALTA COLUMNA 'TIPO DE MEDIDA'")
                # Si falta la columna, no podemos validar el tipo
                # Retornar error inmediatamente
                return (False, " | ".join(errores))
            
            tipo_medida = str(fila.get("TIPO DE MEDIDA", "")).strip().upper()
            
            if not tipo_medida or tipo_medida == "NAN":
                errores.append("FALTA TIPO DE MEDIDA")
                return (False, " | ".join(errores))
            
            # Identificar el tipo (IEI o DSE)
            if "IEI" in tipo_medida:
                medida_tipo = "IEI"
            elif "DSE" in tipo_medida:
                medida_tipo = "DSE"
            else:
                errores.append(f"TIPO DE MEDIDA NO VÁLIDO: {tipo_medida}")
                return (False, " | ".join(errores))
            
            # ============================================================
            # 3. VALIDAR CAMPOS SEGÚN TIPO
            # ============================================================
            if medida_tipo == "IEI":
                # Validar INTERVENTOR
                if "INTERVENTOR" not in expedientes.columns:
                    errores.append("FALTA COLUMNA 'INTERVENTOR'")
                else:
                    interventor = str(fila.get("INTERVENTOR", "")).strip()
                    if not interventor or interventor.upper() == "NAN":
                        errores.append("FALTA INTERVENTOR")
                
                # Validar PLAZO
                if "PLAZO" not in expedientes.columns:
                    errores.append("FALTA COLUMNA 'PLAZO'")
                else:
                    plazo = str(fila.get("PLAZO", "")).strip()
                    if not plazo or plazo.upper() == "NAN":
                        errores.append("FALTA PLAZO")
            
            elif medida_tipo == "DSE":
                # Validar MONTO
                if "MONTO" not in expedientes.columns:
                    errores.append("FALTA COLUMNA 'MONTO'")
                else:
                    monto = str(fila.get("MONTO", "")).strip()
                    if not monto or monto.upper() == "NAN":
                        errores.append("FALTA MONTO")
            
            # ============================================================
            # Retornar resultado
            # ============================================================
            if errores:
                return (False, " | ".join(errores))
            else:
                return (True, "")
        
        except Exception as e:
            logger.error(f"Error en validate_expediente_row (fila {row_idx + 1}): {str(e)}")
            return (False, f"ERROR EN VALIDACIÓN: {str(e)}")
    
    def detect_most_used_type(self, expedientes):
        """
        Detecta cuál es la DEPENDENCIA MÁS USADA en todos los expedientes del Excel.
        
        IMPORTANTE CORRECCIÓN:
        - La comparación NO es por tipo (IEI vs DSE)
        - La comparación es por DEPENDENCIA (21 vs 23)
        - IEI puede estar tanto en dependencia 21 como en 23
        - DSE puede estar tanto en dependencia 21 como en 23
        
        Lógica de decisión:
        1. Si hay más expedientes con dependencia 21Retorna "21"
        2. Si hay más expedientes con dependencia 23Retorna "23"
        3. Si hay EMPATE de dependenciasUsa la DEPENDENCIA del PRIMER expediente como desempate
           - Si primer expediente es dependencia 21Retorna "21"
           - Si primer expediente es dependencia 23Retorna "23"
        
        Args:
            expedientes: DataFrame del Excel
            
        Retorna:
            - "21" si hay más expedientes con dependencia 21, o empate y primer expediente es dependencia 21
            - "23" si hay más expedientes con dependencia 23, o empate y primer expediente es dependencia 23
        """
        try:
            if "DEPENDENCIA" not in expedientes.columns:
                logger.warning("Columna DEPENDENCIA no encontrada")
                return "21"
            
            cuenta_dep21 = 0
            cuenta_dep23 = 0
            
            # Contar dependencias en TODOS los expedientes
            for idx, fila in expedientes.iterrows():
                dependencia = str(fila.get("DEPENDENCIA", "")).strip()
                
                if "21" in dependencia:
                    cuenta_dep21 += 1
                elif "23" in dependencia:
                    cuenta_dep23 += 1
            
            # Loguear resultados
            total = cuenta_dep21 + cuenta_dep23
            if total == 0:
                logger.warning("No se encontraron dependencias válidas en el Excel")
                return "21"
            
            logger.info(f"\n Análisis de dependencias en {total} expedientes:")
            logger.info(f"   Dependencia 21 (PRICO): {cuenta_dep21} ({100*cuenta_dep21/total:.1f}%)")
            logger.info(f"   Dependencia 23 (MEPECO): {cuenta_dep23} ({100*cuenta_dep23/total:.1f}%)")
            
            # ================================================================
            # Determinar la dependencia más usada
            # ================================================================
            if cuenta_dep21 > cuenta_dep23:
                logger.info(f"Dependencia dominante: 21 (PRICO) con {cuenta_dep21} expedientes")
                return "21"
            elif cuenta_dep23 > cuenta_dep21:
                logger.info(f"Dependencia dominante: 23 (MEPECO) con {cuenta_dep23} expedientes")
                return "23"
            else:
                # EMPATE DE DEPENDENCIAS: Usar dependencia del PRIMER expediente como desempate
                logger.info(f"EMPATE de dependencias ({cuenta_dep21} vs {cuenta_dep23})")
                logger.info(f" Usando dependencia del PRIMER expediente como criterio de desempate:")
                
                # Obtener dependencia del primer expediente
                primera_dependencia = str(expedientes.iloc[0].get("DEPENDENCIA", "")).strip()
                logger.info(f"  Primer expediente: Dependencia {primera_dependencia}")
                
                if "21" in primera_dependencia:
                    logger.info(f"Desempate resuelto: 21 (PRICO) - dependencia del primer expediente")
                    return "21"
                elif "23" in primera_dependencia:
                    logger.info(f"Desempate resuelto: 23 (MEPECO) - dependencia del primer expediente")
                    return "23"
                else:
                    logger.warning(f" Dependencia no reconocida: '{primera_dependencia}'. Usando default: 21")
                    return "21"
        
        except Exception as e:
            logger.error(f"Error detectando dependencia más usada: {str(e)}")
            return "21"
    
    def validate_excel_columns(self, expedientes):
        """
        Valida que el Excel tenga TODAS las columnas obligatorias.
        
        CAMBIO IMPORTANTE: Ahora IEI y DSE pueden aparecer en AMBAS dependencias (21 y 23).
        Por lo tanto, se requieren TODAS las columnas:
        - TIPO DE MEDIDA: Obligatorio (detecta IEI o DSE por expediente)
        - INTERVENTOR: Obligatorio (para expedientes IEI)
        - PLAZO: Obligatorio (para expedientes IEI)
        - MONTO: Obligatorio (para expedientes DSE)
        
        Retorna:
            - (True, bucle_inicial) si las columnas son válidas (bucle_inicial puede ser "21" o "23")
            - (False, mensaje_error) si faltan columnas
        """
        try:
            logger.info("\n" + "=" * 70)
            logger.info("VALIDANDO COLUMNAS OBLIGATORIAS DEL EXCEL")
            logger.info("=" * 70)
            
            # Columnas OBLIGATORIAS para soportar tanto IEI como DSE
            columnas_requeridas = ["TIPO DE MEDIDA", "INTERVENTOR", "PLAZO", "MONTO"]
            columnas_faltantes = [col for col in columnas_requeridas if col not in expedientes.columns]
            
            if columnas_faltantes:
                logger.error(f" FALTAN COLUMNAS OBLIGATORIAS: {', '.join(columnas_faltantes)}")
                logger.info("El Excel debe contener TODAS estas columnas:")
                logger.info("   TIPO DE MEDIDA (IEI o DSE)")
                logger.info("   INTERVENTOR (para expedientes IEI)")
                logger.info("   PLAZO (para expedientes IEI)")
                logger.info("   MONTO (para expedientes DSE)")
                return (False, f"Faltan columnas: {', '.join(columnas_faltantes)}")
            
            logger.info(" TODAS LAS COLUMNAS OBLIGATORIAS ESTÁN PRESENTES")
            logger.info("Columnas validadas:")
            logger.info("   TIPO DE MEDIDA")
            logger.info("   INTERVENTOR")
            logger.info("   PLAZO")
            logger.info("   MONTO")
            
            # ================================================================
            # DETECTAR DEPENDENCIA MÁS USADA PARA DECIDIR CUÁL BUCLE COMIENZA
            # ================================================================
            logger.info("\n" + "=" * 70)
            logger.info("DETECTANDO DEPENDENCIA MÁS USADA PARA DECIDIR BUCLE INICIAL")
            logger.info("=" * 70)
            
            bucle_inicial = self.detect_most_used_type(expedientes)
            
            if bucle_inicial == "21":
                logger.info(f" Bucle inicial: 21 (PRICO) - Más expedientes con dependencia 21")
            else:
                logger.info(f" Bucle inicial: 23 (MEPECO) - Más expedientes con dependencia 23")
            
            logger.info("NOTA: Ambas dependencias pueden contener IEI, DSE o ambos tipos")
            logger.info("El tipo de cada expediente será detectado dinámicamente\n")
            
            return (True, bucle_inicial)
        
        except Exception as e:
            logger.error(f"Error en validate_excel_columns: {str(e)}")
            return (False, f"Error validando columnas: {str(e)}")
    
    def load_credentials(self):
        """Carga contraseña desde archivo y lee Excel para determinar dependencia"""
        try:
            # Cargar contraseña
            password_file = SCRIPT_DIR / "contrasena.txt"
            if not password_file.exists():
                logger.error(f"Archivo de contraseña no encontrado: {password_file}")
                return False
            
            with open(password_file, "r", encoding="utf-8") as f:
                self.password = f.read().strip()
            
            logger.info("Contraseña cargada correctamente")
            
            # Cargar Excel - especificar dtype para columns numéricas como string
            excel_file = SCRIPT_DIR / "EXPEDIENTES.xlsx"
            if not excel_file.exists():
                logger.error(f"Archivo Excel no encontrado: {excel_file}")
                return False
            
            # Leer con dtype object para preservar formato original
            expedientes = pd.read_excel(excel_file, engine="openpyxl", dtype=str)
            
            if "DEPENDENCIA" not in expedientes.columns:
                logger.error("El Excel no contiene la columna 'DEPENDENCIA'")
                return False
            
            # ============================================================
            # VALIDAR COLUMNAS OBLIGATORIAS SEGÚN TIPO DE MEDIDA
            # ============================================================
            columnas_validas, tipo_o_error = self.validate_excel_columns(expedientes)
            
            if not columnas_validas:
                logger.error(f" VALIDACIÓN FALLIDA: {tipo_o_error}")
                logger.error("Por favor, verifica que el Excel tenga las columnas obligatorias")
                return False
            
            # Asignar tipo detectado por TIPO DE MEDIDA
            self.dep_type = tipo_o_error
            
            # Mapear dependencia por tipo detectado
            # IMPORTANTE: Ahora IEI y DSE pueden aparecer en AMBAS dependencias (21 y 23)
            # Por lo tanto, solo asignamos la dependencia según el tipo inicial
            if self.dep_type == "21":
                self.dependencia = "0021 I.R. Lima - PRICO"
                logger.info(" Dependencia inicial: 21 (PRICO)")
                logger.info("IMPORTANTE: Esta dependencia puede contener IEI, DSE o ambos")
                logger.info("El tipo se detectará dinámicamente por cada expediente")
            elif self.dep_type == "23":
                self.dependencia = "0023 I.R. Lima - MEPECO"
                logger.info(" Dependencia inicial: 23 (MEPECO)")
                logger.info("IMPORTANTE: Esta dependencia puede contener IEI, DSE o ambos")
                logger.info("El tipo se detectará dinámicamente por cada expediente")
            
            # Extraer expediente (ya es texto en Excel, solo hacer strip)
            if "EXPEDIENTE" in expedientes.columns:
                self.expediente = str(expedientes.iloc[0]["EXPEDIENTE"]).strip()
                logger.info(f"Expediente: {self.expediente} (formato texto)")
            
            logger.info(f"Dependencia: {self.dependencia}")
            return True
        
        except Exception as e:
            logger.error(f"Error cargando credenciales: {str(e)}")
            return False
    
    def wait_for_login_window(self, timeout=30):
        """Espera a que aparezca la ventana de login"""
        try:
            logger.info("Esperando ventana de login...")
            desktop = Desktop(backend="uia")
            end_time = time.time() + timeout
            
            while time.time() < end_time:
                try:
                    win = desktop.window(title="SIRAT")
                    if win.exists(timeout=1):
                        handle = win.handle
                        app = Application(backend="uia").connect(handle=handle)
                        logger.info("Ventana de login encontrada")
                        return app, app.window(handle=handle)
                except Exception:
                    pass
                
                try:
                    win = desktop.window(title_re=".*SIRAT.*")
                    if win.exists(timeout=1):
                        handle = win.handle
                        app = Application(backend="uia").connect(handle=handle)
                        logger.info("Ventana de login encontrada")
                        return app, app.window(handle=handle)
                except Exception:
                    pass
                
                time.sleep(0.5)
            
            logger.error("Timeout esperando ventana de login")
            return None, None
        
        except Exception as e:
            logger.error(f"Error esperando login: {str(e)}")
            return None, None
    
    def detect_password_error(self, timeout=2):
        """
        Detecta si hay un mensaje de error de contraseña incorrecta usando MSAA.
        Busca controles de tipo Text que contengan palabras clave de error.
        Sin OCR ni coordenadas - solo inspección de elementos de la UI.
        
        Retorna:
            - Tupla (True, mensaje) si se detecta error
            - Tupla (False, "") si no hay error
        """
        try:
            logger.info("Verificando si hay mensaje de error...")
            desktop = Desktop(backend="uia")
            end_time = time.time() + timeout
            
            while time.time() < end_time:
                try:
                    # Obtener todas las ventanas visibles
                    all_windows = desktop.windows()
                    
                    for win in all_windows:
                        try:
                            # Obtener todos los descendientes
                            all_descendants = list(win.descendants())
                            
                            # Buscar controles de tipo Text con palabras clave
                            for desc in all_descendants:
                                try:
                                    # Obtener control_type y texto
                                    control_type = None
                                    texto = ""
                                    
                                    try:
                                        if hasattr(desc, 'element_info'):
                                            control_type = desc.element_info.control_type
                                    except:
                                        pass
                                    
                                    try:
                                        texto = desc.window_text()
                                    except:
                                        pass
                                    
                                    # Verificar si es un control Text con contenido
                                    if control_type == "Text" and texto and len(texto.strip()) > 0:
                                        # Buscar palabras clave de error
                                        texto_lower = texto.lower()
                                        palabras_clave = [
                                            "no puede ser accedido",
                                            "estimado usuario",
                                            "aplicativo"
                                        ]
                                        
                                        if any(palabra in texto_lower for palabra in palabras_clave):
                                            logger.error(f"Mensaje de error detectado: {texto}")
                                            return (True, texto)
                                
                                except:
                                    pass
                        
                        except:
                            pass
                
                except:
                    pass
                
                time.sleep(0.3)
            
            return (False, "")
        
        except Exception as e:
            logger.error(f"Error en detect_password_error: {e}")
            return (False, "")
    
    def detect_monto_aviso(self, timeout=2):
        """
        Detecta y extrae el mensaje de aviso que aparece después de presionar ALT+A en DSE.
        Puede variar dependiendo del monto ingresado (ej: "El monto ingresado excede en más del 20% el Saldo del Expediente").
        
        Usa MSAA para inspeccionar controles de tipo Text sin OCR ni coordenadas.
        
        Retorna:
            - Tupla (True, mensaje) si se detecta un mensaje de aviso
            - Tupla (False, "") si no hay mensaje
        """
        try:
            logger.info("Verificando si hay mensaje de aviso del MONTO...")
            desktop = Desktop(backend="uia")
            end_time = time.time() + timeout
            
            while time.time() < end_time:
                try:
                    # Obtener todas las ventanas visibles
                    all_windows = desktop.windows()
                    
                    for win in all_windows:
                        try:
                            # Obtener todos los descendientes
                            all_descendants = list(win.descendants())
                            
                            # Buscar controles de tipo Text que contengan mensajes
                            for desc in all_descendants:
                                try:
                                    # Obtener control_type y texto
                                    control_type = None
                                    texto = ""
                                    
                                    try:
                                        if hasattr(desc, 'element_info'):
                                            control_type = desc.element_info.control_type
                                    except:
                                        pass
                                    
                                    try:
                                        texto = desc.window_text()
                                    except:
                                        pass
                                    
                                    # Verificar si es un control Text con contenido y es un mensaje de aviso
                                    if control_type == "Text" and texto and len(texto.strip()) > 0:
                                        # Buscar palabras clave típicas de mensajes de aviso/error
                                        texto_lower = texto.lower()
                                        palabras_clave = [
                                            "monto",
                                            "excede",
                                            "saldo",
                                            "expediente",
                                            "aviso",
                                            "error"
                                        ]
                                        
                                        if any(palabra in texto_lower for palabra in palabras_clave):
                                            logger.info(f"Mensaje de aviso detectado: {texto}")
                                            return (True, texto)
                                
                                except:
                                    pass
                        
                        except:
                            pass
                
                except:
                    pass
                
                time.sleep(0.3)
            
            return (False, "")
        
        except Exception as e:
            logger.error(f"Error en detect_monto_aviso: {e}")
            return (False, "")
    
    def detect_expediente_error(self, timeout=2):
        """
        Detecta si hay un mensaje de error de expediente inválido usando MSAA.
        Busca controles de tipo Text que contengan palabras clave de error.
        
        Mensaje esperado: "El número de Expediente Coactivo ingresado no es válido"
        
        Sin OCR ni coordenadas - solo inspección de elementos de la UI.
        
        Retorna:
            - Tupla (True, mensaje) si se detecta error
            - Tupla (False, "") si no hay error
        """
        try:
            logger.info("Verificando si hay mensaje de error de expediente...")
            desktop = Desktop(backend="uia")
            end_time = time.time() + timeout
            
            while time.time() < end_time:
                try:
                    # Obtener todas las ventanas visibles
                    all_windows = desktop.windows()
                    
                    for win in all_windows:
                        try:
                            # Obtener todos los descendientes
                            all_descendants = list(win.descendants())
                            
                            # Buscar controles de tipo Text que contengan mensajes
                            for desc in all_descendants:
                                try:
                                    # Obtener control_type y texto
                                    control_type = None
                                    texto = ""
                                    
                                    try:
                                        if hasattr(desc, 'element_info'):
                                            control_type = desc.element_info.control_type
                                    except:
                                        pass
                                    
                                    try:
                                        texto = desc.window_text()
                                    except:
                                        pass
                                    
                                    # Verificar si es un control Text con contenido relacionado a expediente
                                    if control_type == "Text" and texto and len(texto.strip()) > 0:
                                        # Buscar palabras clave típicas de error de expediente
                                        texto_lower = texto.lower()
                                        palabras_clave = [
                                            "expediente",
                                            "ingresado",
                                            "válido",
                                            "no es válido",
                                            "coactivo"
                                        ]
                                        
                                        # Búsqueda específica del patrón
                                        if ("expediente" in texto_lower and "válido" in texto_lower) or \
                                           ("expediente" in texto_lower and "ingresado" in texto_lower):
                                            logger.warning(f"Error de expediente detectado: {texto}")
                                            return (True, texto)
                                
                                except:
                                    pass
                        
                        except:
                            pass
                
                except:
                    pass
                
                time.sleep(0.3)
            
            return (False, "")
        
        except Exception as e:
            logger.error(f"Error en detect_expediente_error: {e}")
            return (False, "")
    
    def detect_expediente_aviso(self, timeout=2):
        """
        Detecta si hay un mensaje de aviso de expediente después de presionar ALT+A en IEI.
        Este aviso es específico de Dependencia 21 (PRICO - Trabar Intervención en Información).
        
        Busca DOS tipos de mensajes posibles:
        1. Si hay embargos activos: "El Expediente..."
        2. Si NO hay embargos: "¿Desea Ud. grabar la Resolución?"
        
        Usa MSAA para inspeccionar controles de tipo Text sin OCR ni coordenadas.
        
        Retorna:
            - Tupla (True, mensaje) si se detecta un mensaje de aviso
            - Tupla (False, "") si no hay mensaje
        """
        try:
            logger.info("Verificando si hay mensaje de aviso de expediente (IEI)...")
            desktop = Desktop(backend="uia")
            end_time = time.time() + timeout
            
            while time.time() < end_time:
                try:
                    # Obtener todas las ventanas visibles
                    all_windows = desktop.windows()
                    
                    for win in all_windows:
                        try:
                            # Obtener todos los descendientes
                            all_descendants = list(win.descendants())
                            
                            # Buscar controles de tipo Text que contengan mensajes
                            for desc in all_descendants:
                                try:
                                    # Obtener control_type y texto
                                    control_type = None
                                    texto = ""
                                    
                                    try:
                                        if hasattr(desc, 'element_info'):
                                            control_type = desc.element_info.control_type
                                    except:
                                        pass
                                    
                                    try:
                                        texto = desc.window_text()
                                    except:
                                        pass
                                    
                                    # Verificar si es un control Text con contenido
                                    if control_type == "Text" and texto and len(texto.strip()) > 0:
                                        # Búsqueda ROBUSTA: múltiples palabras clave
                                        texto_lower = texto.lower()
                                        
                                        # Patrón: "El Expediente XXX correspondiente al RUC XXX tiene X Embargos activos..."
                                        # Búsqueda robusta: "expediente" + "ruc" + "embargo" + "activo"
                                        if ("expediente" in texto_lower and 
                                            "ruc" in texto_lower and 
                                            "embargo" in texto_lower and 
                                            "activo" in texto_lower):
                                            logger.warning(f"Aviso de embargos activos detectado")
                                            return (True, texto)
                                
                                except:
                                    pass
                        
                        except:
                            pass
                
                except:
                    pass
                
                time.sleep(0.3)
            
            return (False, "")
        
        except Exception as e:
            logger.error(f"Error en detect_expediente_aviso: {e}")
            return (False, "")
    
    def extract_ruc_from_message(self, mensaje):
        """
        Extrae el número RUC del mensaje de aviso de expediente.
        
        Ejemplo: "El Expediente XXX correspondiente al RUC 12345678910..."
        Retorna: "12345678910"
        
        Args:
            mensaje: Texto del mensaje que contiene el RUC
            
        Retorna:
            - String con el número RUC si se encuentra
            - String vacío "" si no se encuentra
        """
        try:
            logger.info("Extrayendo RUC del mensaje...")
            
            # Buscar "RUC" en el mensaje (case-insensitive)
            ruc_index = mensaje.upper().find("RUC")
            
            if ruc_index == -1:
                logger.warning("No se encontró 'RUC' en el mensaje")
                return ""
            
            # Obtener el texto después de "RUC"
            texto_despues_ruc = mensaje[ruc_index + 3:].strip()
            
            logger.info(f"Texto después de 'RUC': '{texto_despues_ruc[:50]}'")
            
            # Extraer los primeros dígitos consecutivos (el RUC)
            ruc = ""
            for caracter in texto_despues_ruc:
                if caracter.isdigit():
                    ruc += caracter
                elif ruc:  # Si ya tenemos dígitos y encontramos un no-dígito, parar
                    break
            
            if ruc:
                logger.info(f" RUC extraído: {ruc}")
                return ruc
            else:
                logger.warning("No se encontraron dígitos después de 'RUC'")
                return ""
        
        except Exception as e:
            logger.error(f"Error extrayendo RUC: {e}")
            return ""
    
    def detect_resolucion_coactiva_aviso(self, timeout=2):
        """
        Detecta el mensaje de confirmación de Resolución Coactiva después de ALT+S en IEI.
        Busca el patrón: "Se grabó la Resolución Coactiva con el número XXXX..."
        
        Usa MSAA para inspeccionar controles de tipo Text sin OCR ni coordenadas.
        
        Retorna:
            - Tupla (True, mensaje) si se detecta el mensaje
            - Tupla (False, "") si no hay mensaje
        """
        try:
            logger.info("Verificando si hay mensaje de Resolución Coactiva...")
            desktop = Desktop(backend="uia")
            end_time = time.time() + timeout
            
            while time.time() < end_time:
                try:
                    all_windows = desktop.windows()
                    
                    for win in all_windows:
                        try:
                            all_descendants = list(win.descendants())
                            
                            for desc in all_descendants:
                                try:
                                    control_type = None
                                    texto = ""
                                    
                                    try:
                                        if hasattr(desc, 'element_info'):
                                            control_type = desc.element_info.control_type
                                    except:
                                        pass
                                    
                                    try:
                                        texto = desc.window_text()
                                    except:
                                        pass
                                    
                                    # Verificar si es un control Text que contiene "Se grabó la Resolución Coactiva"
                                    if control_type == "Text" and texto and len(texto.strip()) > 0:
                                        # Buscar patrón "Se grabó la Resolución Coactiva" (case-insensitive)
                                        if "se grabó" in texto.lower() and "resolución coactiva" in texto.lower():
                                            logger.warning(f" MENSAJE DE RESOLUCIÓN COACTIVA DETECTADO: {texto[:100]}...")
                                            return (True, texto)
                                
                                except:
                                    pass
                        
                        except:
                            pass
                
                except:
                    pass
                
                time.sleep(0.3)
            
            return (False, "")
        
        except Exception as e:
            logger.error(f"Error en detect_resolucion_coactiva_aviso: {e}")
            return (False, "")
    
    def extract_resolucion_coactiva_number(self, mensaje):
        """
        Extrae el número de Resolución Coactiva del mensaje.
        
        Ejemplo: "Se grabó la Resolución Coactiva con el número 0290079364147"
        Retorna: "0290079364147"
        
        Args:
            mensaje: Texto del mensaje que contiene el número de RC
            
        Retorna:
            - String con el número de RC si se encuentra (preservando 0 inicial)
            - String vacío "" si no se encuentra
        """
        try:
            logger.info("Extrayendo número de Resolución Coactiva del mensaje...")
            
            # Buscar "número" en el mensaje (case-insensitive)
            numero_index = mensaje.lower().find("número")
            
            if numero_index == -1:
                logger.warning("No se encontró 'palabra 'número' en el mensaje")
                return ""
            
            # Obtener el texto después de "número"
            texto_despues_numero = mensaje[numero_index + 6:].strip()
            
            logger.info(f"Texto después de 'número': '{texto_despues_numero[:50]}'")
            
            # Extraer los primeros dígitos consecutivos (el número de RC)
            rc_number = ""
            for caracter in texto_despues_numero:
                if caracter.isdigit():
                    rc_number += caracter
                elif rc_number:  # Si ya se encontraron dígitos y ahora hay no-dígito, parar
                    break
            
            if rc_number:
                logger.info(f" Número de RC extraído: {rc_number}")
                return rc_number
            else:
                logger.warning("No se encontraron dígitos después de 'número'")
                return ""
        
        except Exception as e:
            logger.error(f"Error extrayendo número de RC: {e}")
            return ""
    
    def login(self):
        try:
            logger.info("\nIniciando proceso de login...")
            
            # Cargar credenciales
            if not self.load_credentials():
                return False
            
            # Esperar ventana de login
            app, dlg = self.wait_for_login_window(timeout=30)
            if dlg is None:
                logger.error("No se pudo encontrar la ventana de login")
                return False
            
            time.sleep(1)
            
            # Buscar campos de entrada
            logger.info("Buscando campos de dependencia y contraseña...")
            try:
                dependencia_edit = dlg.child_window(auto_id="1001", control_type="Edit")
                password_edit = dlg.child_window(auto_id="1005", control_type="Edit")
                logger.info("Campos encontrados por auto_id")
            except Exception as e:
                logger.warning(f"No se encontraron campos por auto_id: {e}")
                try:
                    edits = dlg.descendants(control_type="Edit")
                    edits_list = list(edits)
                    dependencia_edit = edits_list[0] if len(edits_list) > 0 else None
                    password_edit = edits_list[1] if len(edits_list) > 1 else None
                except Exception:
                    dependencia_edit = None
                    password_edit = None
            
            if dependencia_edit is None or password_edit is None:
                logger.error("No se pudieron localizar los campos de login")
                return False
            
            # Ingresar dependencia
            logger.info(f"Ingresando dependencia: {self.dependencia}")
            dependencia_edit.set_focus()
            time.sleep(0.3)
            pyautogui.write(self.dependencia, interval=0.05)
            time.sleep(0.5)
            
            # Ingresar contraseña
            logger.info("Ingresando contraseña...")
            password_edit.set_focus()
            time.sleep(0.3)
            
            # Intentar usar write() primero (para contraseñas simples)
            try:
                logger.info("Ingresando contraseña con write()...")
                pyautogui.write(self.password, interval=0.05)
                logger.info(" write() exitoso")
            except Exception as e:
                # Si falla (ej: caracteres especiales), usar typewrite() carácter por carácter
                logger.info(f"write() falló: {e}")
                logger.info("Fallback: Intentando con typewrite() carácter por carácter...")
                
                # Limpiar campo primero
                pyautogui.hotkey('ctrl', 'a')
                time.sleep(0.1)
                pyautogui.press('delete')
                time.sleep(0.2)
                
                # Escribir carácter por carácter
                for i, char in enumerate(self.password):
                    logger.info(f"  [{i+1}/{len(self.password)}] Escribiendo '{char}'...")
                    pyautogui.typewrite(char)
                    time.sleep(0.05)
                
                logger.info(" typewrite() completado")
            
            time.sleep(0.5)
            
            # Hacer clic en Aceptar
            logger.info("Haciendo clic en Aceptar...")
            try:
                aceptar_btn = dlg.child_window(title="Aceptar", control_type="Button")
                aceptar_btn.invoke()
            except Exception:
                logger.warning("No se encontró botón 'Aceptar', intentando con Enter...")
                pyautogui.press('return')
            
            # Pequeña pausa para que aparezca el posible error
            time.sleep(1)
            
            # Verificar si hay mensaje de error de contraseña
            error_detected, error_message = self.detect_password_error(timeout=2)
            if error_detected:
                logger.error(f"Contraseña incorrecta. Mensaje: {error_message}")
                
                # Presionar ENTER
                logger.info("Presionando ENTER...")
                time.sleep(0.5)
                pyautogui.press('return')
                time.sleep(0.5)
                
                # Presionar ALT+C para cerrar RSIRAT
                logger.info("Presionando ALT+C para cerrar RSIRAT...")
                time.sleep(0.5)
                pyautogui.hotkey('alt', 'c')
                time.sleep(1)
                
                # Actualizar Excel con el resultado
                self.update_excel_result("CONTRASEÑA INCORRECTA")
                
                logger.info("Proceso finalizado")
                return False
            
            logger.info("Login completado")
            
            # Esperar dinámicamente por el menú
            menu_detected = False
            start_wait = time.time()
            max_wait = 1
            
            while time.time() - start_wait < max_wait:
                try:
                    menu_window = app.window(title_re=".*Menú.*")
                    if menu_window and menu_window.is_visible():
                        menu_detected = True
                        logger.info("Menú de opciones detectado")
                        time.sleep(2)
                        break
                except Exception:
                    pass
                
                time.sleep(0.5)
            
            if not menu_detected:
                logger.warning("Menú de opciones no detectado, continuando...")
            
            return True
        
        except Exception as e:
            logger.error(f"Error en login: {str(e)}")
            return False
    
    def open_application(self):
        """Abre RSIRAT usando el acceso directo"""
        logger.info("=" * 70)
        logger.info("INICIANDO AUTOMATIZACIÓN DE RSIRAT (32-BIT)")
        logger.info("=" * 70)
        logger.info(f"Buscando acceso directo en: {SHORTCUT_PATH}")
        
        if not SHORTCUT_PATH.exists():
            logger.error(f"Acceso directo no encontrado: {SHORTCUT_PATH}")
            return False
        
        try:
            logger.info(f"Abriendo: {SHORTCUT_PATH.name}")
            os.startfile(str(SHORTCUT_PATH))
            time.sleep(4)
            logger.info("Aplicación abierta correctamente")
            return True
        except Exception as e:
            logger.error(f"Error al abrir aplicación: {str(e)}")
            return False
    
    def update_excel_result(self, resultado):
        """
        Actualiza el archivo R_EXPEDIENTES.xlsx con el resultado del proceso.
        Este es un nuevo archivo de resultados, NO modifica el Excel original.
        
        Escribe en la fila correspondiente al expediente procesado (self.primer_expediente_idx).
        
        Args:
            resultado: El texto a escribir en la columna RESULTADO
        """
        try:
            logger.info(f"Actualizando R_EXPEDIENTES.xlsx con resultado: {resultado}")
            
            excel_file = SCRIPT_DIR / "EXPEDIENTES.xlsx"
            resultado_file = SCRIPT_DIR / "R_EXPEDIENTES.xlsx"
            
            if not excel_file.exists():
                logger.error(f"Archivo Excel no encontrado: {excel_file}")
                return False
            
            # Si R_EXPEDIENTES.xlsx no existe, crear una copia del original
            from openpyxl import load_workbook
            from openpyxl.styles import numbers
            
            if not resultado_file.exists():
                logger.info("R_EXPEDIENTES.xlsx no existe, creando copia del Excel original...")
                import shutil
                shutil.copy(excel_file, resultado_file)
            
            # Abrir R_EXPEDIENTES.xlsx
            wb = load_workbook(resultado_file)
            ws = wb.active
            
            # Encontrar la columna RESULTADO o crearla
            headers = {}
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value:
                    headers[cell.value] = col_idx
            
            # Si no existe la columna RESULTADO, crearla
            if "RESULTADO" not in headers:
                resultado_col = len(headers) + 1
                ws.cell(row=1, column=resultado_col, value="RESULTADO")
            else:
                resultado_col = headers["RESULTADO"]
            
            # Escribir el resultado en la fila correcta (basada en primer_expediente_idx)
            # primer_expediente_idx es 0-based, así que row = idx + 2 (fila 1 = header, fila 2 = idx 0)
            target_row = self.primer_expediente_idx + 2
            logger.info(f"Escribiendo resultado en fila {target_row} (idx={self.primer_expediente_idx})")
            celda_resultado = ws.cell(row=target_row, column=resultado_col, value=resultado)
            
            # Forzar formato de texto para preservar el 0 inicial
            if isinstance(resultado, str) and resultado.isdigit():
                celda_resultado.number_format = '@'
            
            # Guardar el archivo
            wb.save(resultado_file)
            logger.info(f"R_EXPEDIENTES.xlsx actualizado correctamente (fila {target_row})")
            # Actualizar estado de completado según resultado escrito
            try:
                if isinstance(resultado, str):
                    res_up = resultado.strip().upper()
                    if resultado.isdigit() or res_up == 'MONTO MAYOR':
                        self.last_exp_completed = True
                    else:
                        self.last_exp_completed = False
                else:
                    self.last_exp_completed = False
            except Exception:
                pass
            return True
        
        except Exception as e:
            logger.error(f"Error actualizando R_EXPEDIENTES.xlsx: {e}")
            return False

    def mark_invalid_expediente_in_results(self, row_idx, motivo="EXP. INVALIDO"):
        """
        Marca un expediente como inválido en R_EXPEDIENTES.xlsx con el motivo específico.
        
        Args:
            row_idx: Índice 0-based de la fila en el Excel
            motivo: Razón por la que es inválido (ej: "FALTA MONTO", "FALTA INTERVENTOR", etc.)
        """
        try:
            logger.info(f"Marcando expediente inválido en fila {row_idx + 2}...")
            logger.info(f"Motivo: {motivo}")
            
            excel_file = SCRIPT_DIR / "EXPEDIENTES.xlsx"
            resultado_file = SCRIPT_DIR / "R_EXPEDIENTES.xlsx"
            
            from openpyxl import load_workbook
            
            # Si R_EXPEDIENTES.xlsx no existe, crear una copia del original
            if not resultado_file.exists():
                logger.info("R_EXPEDIENTES.xlsx no existe, creando copia del Excel original...")
                import shutil
                shutil.copy(excel_file, resultado_file)
            
            # Abrir R_EXPEDIENTES.xlsx
            wb = load_workbook(resultado_file)
            ws = wb.active
            
            # Encontrar la columna RESULTADO o crearla
            headers = {}
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value:
                    headers[cell.value] = col_idx
            
            # Si no existe la columna RESULTADO, crearla
            if "RESULTADO" not in headers:
                resultado_col = len(headers) + 1
                ws.cell(row=1, column=resultado_col, value="RESULTADO")
            else:
                resultado_col = headers["RESULTADO"]
            
            # Escribir el motivo en la fila especificada
            target_row = row_idx + 2
            ws.cell(row=target_row, column=resultado_col, value=motivo)
            
            # Guardar el archivo
            wb.save(resultado_file)
            logger.info(f" Expediente marcado como inválido en R_EXPEDIENTES.xlsx (fila {target_row}): {motivo}")
            # Marcar que el expediente NO quedó completado
            try:
                self.last_exp_completed = False
            except Exception:
                pass
            return True
        
        except Exception as e:
            logger.error(f"Error marcando expediente inválido: {e}")
            return False
    
    def click_cobranza_coactiva(self):
        """
        ENCADENA TODOS LOS PASOS DESDE CERO para procesar UN expediente:
        
        1. Click en "Cobranza Coactiva"
        2. 4 clics en "Exp. Cob. Coactiva - Individual"
        3. Ingresa y valida expediente
        4. Presiona ALT+A para validar ejecutor
        5. Ejecuta flujo post-embargo (IEI o DSE)
        
        Este método se llama para CADA expediente en un lote.
        Después de ALT+F4 (cierre de app), esta secuencia se repite desde cero.
        """
        logger.info("\n" + "=" * 70)
        logger.info("INICIANDO SECUENCIA COMPLETA: COBRANZA COACTIVA → EMBARGO")
        logger.info("=" * 70)
        
        try:
            # PASO 1: Buscar y clicar "Cobranza Coactiva"
            logger.info("\nPASO 1: Buscando 'Cobranza Coactiva'...")
            if not self._click_cobranza_coactiva_element():
                logger.error("No se pudo encontrar 'Cobranza Coactiva'")
                return False
            
            time.sleep(1)
            
            # PASO 2: Hacer 4 clics en "Exp. Cob. Coactiva - Individual"
            logger.info("\nPASO 2: Haciendo 4 clics en 'Exp. Cob. Coactiva - Individual'...")
            if not self.click_exp_cob_individual():
                logger.error("Error en 4 clics")
                return False
            
            time.sleep(0.5)
            
            # PASO 3: Ingresar y validar expediente
            logger.info("\nPASO 3: Ingresando y validando expediente...")
            if not self.enter_expediente_field():
                logger.error("Error validando expediente")
                return False
            
            time.sleep(0.5)
            
            # PASO 4: ALT+A para validar ejecutor
            logger.info("\nPASO 4: Validando ejecutor con ALT+A...")
            if not self.validate_executor():
                logger.error("Error en validación de ejecutor")
                return False
            
            time.sleep(1)
            
            # PASO 5: Procesar embargo según tipo (IEI o DSE)
            logger.info("\nPASO 5: Procesando embargo según tipo de medida...")
            if not self.handle_post_embargo_flow():
                logger.error("Error en flujo post-embargo")
                return False
            
            logger.info("\n✓ Expediente procesado completamente")
            return True
        
        except Exception as e:
            logger.error(f"Error en click_cobranza_coactiva: {str(e)}")
            return False
    
    def _click_cobranza_coactiva_element(self):
        """
        Búsqueda INTERNA de "Cobranza Coactiva" - Solo busca y clica el elemento.
        No encadena pasos siguientes (eso lo hace click_cobranza_coactiva).
        """
        logger.info("Buscando elemento 'Cobranza Coactiva' en menú...")
        
        try:
            desktop = Desktop(backend="uia")
            
            # Buscar ventana del menú
            try:
                win = desktop.window(title_re=".*Menú.*")
                if win.exists(timeout=1):
                    logger.info("Ventana de menú encontrada")
                    app_window = win
                else:
                    raise Exception("Ventana de menú no encontrada")
            except:
                try:
                    win = desktop.window(title_re=".*SIRAT.*")
                    if win.exists(timeout=1):
                        logger.info("Ventana SIRAT encontrada")
                        app_window = win
                    else:
                        raise Exception("Ventana SIRAT no encontrada")
                except:
                    logger.error("No se encontraron ventanas de menú o SIRAT")
                    return False
            
            # Buscar "Cobranza Coactiva" en descendientes
            try:
                descendants = list(app_window.descendants())
                
                for descendant in descendants:
                    try:
                        window_text = descendant.window_text().strip()
                        
                        if window_text == "Cobranza Coactiva":
                            logger.info(f"✓ 'Cobranza Coactiva' encontrado")
                            rect = descendant.rectangle()
                            
                            if rect.left > 0 or rect.top > 0:
                                center_x = (rect.left + rect.right) // 2
                                center_y = (rect.top + rect.bottom) // 2
                                logger.info(f"Haciendo clic en ({center_x}, {center_y})")
                                pyautogui.click(center_x, center_y)
                                time.sleep(1)
                                logger.info(" Clic completado")
                                return True
                    except:
                        pass
            except:
                pass
            
            logger.warning("'Cobranza Coactiva' no encontrado")
            return False
        
        except Exception as e:
            logger.error(f"Error en _click_cobranza_coactiva_element: {str(e)}")
            return False
    
    def click_exp_cob_individual(self):
        """
        Hace 4 clics en 'Exp. Cob. Coactiva - Individual' usando MSAA.
        Este elemento aparece después de clicar en 'Cobranza Coactiva'.
        """
        logger.info("Buscando 'Exp. Cob. Coactiva - Individual' para hacer 4 clics...")
        
        try:
            desktop = Desktop(backend="uia")
            
            # Buscar ventana de menú o principal
            menu_windows = []
            try:
                win = desktop.window(title_re=".*Menú.*")
                if win.exists(timeout=1):
                    menu_windows.append(win)
                    logger.info(f"Ventana de menú encontrada: {win.window_text()}")
            except Exception:
                pass
            
            # Si no se encontró ventana de menú, buscar ventana SIRAT general
            if not menu_windows:
                try:
                    win = desktop.window(title_re=".*SIRAT.*")
                    if win.exists(timeout=1):
                        menu_windows.append(win)
                        logger.info(f"Ventana SIRAT encontrada: {win.window_text()}")
                except Exception:
                    pass
            
            if not menu_windows:
                logger.error("No se encontraron ventanas de SIRAT o Menú")
                return False
            
            app_window = menu_windows[0]
            
            # Buscar el elemento "Exp. Cob. Coactiva - Individual"
            logger.info("Buscando elemento en descendientes...")
            
            try:
                # Intentar encontrar por titulo exacto primero
                exp_control = app_window.child_window(title="Exp. Cob. Coactiva - Individual")
                if exp_control.exists(timeout=2):
                    logger.info("Control 'Exp. Cob. Coactiva - Individual' encontrado por titulo exacto")
                    for i in range(4):
                        try:
                            # Obtener coordenadas del control
                            rect = exp_control.rectangle()
                            center_x = (rect.left + rect.right) // 2
                            center_y = (rect.top + rect.bottom) // 2
                            logger.info(f"Clic {i+1}/4 en coordenadas: ({center_x}, {center_y})")
                            pyautogui.click(center_x, center_y)
                            time.sleep(0.3)
                        except Exception as e:
                            logger.warning(f"Error en clic {i+1}/4: {e}")
                            time.sleep(0.3)
                    logger.info("4 clics completados exitosamente")
                    time.sleep(0.5)
                    
                    # Ahora ingresar el expediente
                    logger.info("Procediendo a ingresar expediente...")
                    return self.enter_expediente_field()
            except Exception as e:
                logger.info(f"No se encontró por titulo exacto: {e}")
            
            # Buscar en descendientes
            try:
                logger.info("Buscando en descendientes del control...")
                descendants = app_window.descendants()
                
                for descendant in descendants:
                    try:
                        # Buscar elemento que contenga "Exp. Cob. Coactiva - Individual"
                        desc_text = descendant.window_text().strip()
                        if "Exp. Cob. Coactiva" in desc_text and "Individual" in desc_text:
                            logger.info(f"Encontrado control: {desc_text}")
                            logger.info(f"Control type: {descendant.element_info.control_type}")
                            
                            # Hacer 4 clics usando coordenadas directas
                            for i in range(4):
                                try:
                                    rect = descendant.rectangle()
                                    center_x = (rect.left + rect.right) // 2
                                    center_y = (rect.top + rect.bottom) // 2
                                    logger.info(f"Clic {i+1}/4 en coordenadas: ({center_x}, {center_y})")
                                    pyautogui.click(center_x, center_y)
                                    time.sleep(0.3)
                                except Exception as err:
                                    logger.warning(f"Error en clic {i+1}/4: {err}")
                                    time.sleep(0.3)
                            
                            logger.info("4 clics completados exitosamente")
                            time.sleep(0.5)
                            
                            # Ahora ingresar el expediente
                            logger.info("Procediendo a ingresar expediente...")
                            return self.enter_expediente_field()
                    except Exception:
                        pass
                
                logger.warning("No se encontró elemento 'Exp. Cob. Coactiva - Individual' en descendientes")
            
            except Exception as e:
                logger.error(f"Error buscando en descendientes: {e}")
            
            return False
        
        except Exception as e:
            logger.error(f"Error en click_exp_cob_individual: {str(e)}")
            return False
    
    def enter_expediente_field(self):
        """
        Busca el campo 'Número' en la ventana y digita expedientes desde el Excel secuencialmente.
        Verifica si cada expediente es válido detectando mensaje de error usando detect_expediente_error().
        
        Flujo:
        1. Digita el expediente
        2. Presiona Enter
        3. Detecta si hay error de expediente (usando MSAA)
        4. Si hay error: 
           - Presiona ENTER para cerrar el diálogo
           - Borra el expediente inválido
           - Registra "EXP. INVALIDO" en Excel
           - Intenta con el siguiente expediente del Excel
        5. Si no hay error: Continúa con validación de ejecutor (ALT+A)
        """
        # Cargar datos del Excel
        excel_file = SCRIPT_DIR / "EXPEDIENTES.xlsx"
        expedientes = pd.read_excel(excel_file, engine="openpyxl", dtype=str)
        
        # Procesar expedientes hasta encontrar uno válido
        for idx in range(len(expedientes)):
            exp_actual = str(expedientes.iloc[idx]["EXPEDIENTE"]).strip()
            
            logger.info(f"\nIntentando expediente {idx + 1} de {len(expedientes)}: {exp_actual}")
            
            try:
                # Digitar el expediente
                logger.info(f"Digitando expediente: '{exp_actual}'")
                pyautogui.write(exp_actual, interval=0.05)
                logger.info(f" Expediente '{exp_actual}' ingresado")
                time.sleep(0.5)
                
                # Presionar Enter para verificar expediente
                logger.info("Presionando Enter para verificar expediente...")
                pyautogui.press('return')
                time.sleep(1)
                
                # Pequeña pausa para que aparezca el posible error
                time.sleep(0.5)
                
                # Verificar si hay mensaje de error de expediente
                error_detected, error_message = self.detect_expediente_error(timeout=2)
                if error_detected:
                    logger.warning(f"✗ Expediente inválido: {error_message}")
                    
                    # Marcar como inválido en R_EXPEDIENTES.xlsx
                    self.mark_invalid_expediente_in_results(idx)
                    
                    # Presionar ENTER para cerrar el diálogo de error
                    logger.info("Presionando ENTER para cerrar el diálogo de error...")
                    pyautogui.press('return')
                    time.sleep(0.5)
                    
                    # Borrar el expediente que quedó en el campo con Ctrl+Backspace x20
                    logger.info("Borrando el expediente del campo con Ctrl+Backspace x20...")
                    for i in range(20):
                        pyautogui.hotkey('ctrl', 'backspace')
                        time.sleep(0.1)
                    time.sleep(0.5)
                    
                    logger.info("Continuando con el siguiente expediente...")
                    continue  # Pasar al siguiente expediente
                
                logger.info(f" Expediente válido: {exp_actual}")
                logger.info(f"Expediente encontrado en posición: {idx} (fila {idx + 2})")
                logger.info("Procediendo a validación de ejecutor...")
                
                # Guardar el índice del expediente válido encontrado para usar en el bucle
                self.primer_expediente_idx = idx
                
                # Si el expediente es válido, proceder a validar ejecutor (ALT+A)
                return self.validate_executor()
            
            except Exception as e:
                logger.error(f"Error procesando expediente {idx + 1}: {str(e)}")
                continue
        
        # Si se alcanza aquí, todos los expedientes fueron inválidos
        logger.error("Todos los expedientes fueron inválidos")
        return False

    def enter_specific_expediente(self, row_idx):
        """
        Ingresa y valida un expediente específico por índice (0-based).
        Retorna True si el expediente fue aceptado en SIRAT (válido),
        False si fue inválido (y marca en R_EXPEDIENTES.xlsx).
        Esta rutina NO itera sobre todo el Excel; solo intenta el row_idx.
        """
        try:
            excel_file = SCRIPT_DIR / "EXPEDIENTES.xlsx"
            expedientes = pd.read_excel(excel_file, engine="openpyxl", dtype=str)

            exp_actual = str(expedientes.iloc[row_idx]["EXPEDIENTE"]).strip()
            logger.info(f"Intentando en campo específico: fila {row_idx + 1}: {exp_actual}")

            # Digitar expediente
            pyautogui.write(exp_actual, interval=0.05)
            time.sleep(0.5)
            pyautogui.press('return')
            time.sleep(1)

            # Verificar si hay mensaje de error
            error_detected, error_message = self.detect_expediente_error(timeout=2)
            if error_detected:
                logger.warning(f"✗ Expediente inválido: {error_message}")
                self.mark_invalid_expediente_in_results(row_idx)
                pyautogui.press('return')
                time.sleep(0.5)
                # limpiar campo
                pyautogui.hotkey('ctrl', 'backspace')
                time.sleep(0.2)
                # No quedó completado
                self.last_exp_completed = False
                return False

            # Expediente válido
            logger.info(f" Expediente válido: {exp_actual}")
            self.primer_expediente_idx = row_idx
            return True

        except Exception as e:
            logger.error(f"Error en enter_specific_expediente fila {row_idx + 1}: {e}")
            return False
    
    def check_expediente_error(self, app_window):
        """
        Verifica si aparece el mensaje de error de expediente inválido.
        Si aparece, actualiza Excel, cierra ventana y retorna False.
        Si no aparece, retorna True (expediente válido).
        """
        logger.info("Verificando si el expediente es válido...")
        time.sleep(0.5)
        
        try:
            # Buscar diálogo de error por ventana
            desktop = Desktop(backend="uia")
            
            # Patrón 1: "Selección de Expediente Coactivo - Error"
            try:
                error_dialog = desktop.window(title_re=".*Expediente.*Error.*")
                if error_dialog.exists(timeout=2):
                    logger.warning("EXPEDIENTE INVÁLIDO! Diálogo de error detectado")
                    
                    # Actualizar Excel marcando como inválido
                    logger.info("Registrando NRO EXPEDIENTE INVALIDO en Excel...")
                    self.update_excel_executor_result("NRO EXPEDIENTE INVALIDO")
                    
                    # Buscar botón "Aceptar" en el diálogo
                    try:
                        logger.info("Buscando botón 'Aceptar' en el diálogo de error...")
                        aceptar_btn = error_dialog.child_window(title="Aceptar", control_type="Button")
                        if aceptar_btn.exists(timeout=1):
                            logger.info("Botón 'Aceptar' encontrado, haciendo clic...")
                            aceptar_btn.invoke()
                            time.sleep(1)
                    except Exception as e:
                        logger.warning(f"No se encontró botón 'Aceptar': {e}")
                        # Fallback: presionar Enter
                        logger.info("Fallback: presionando Enter...")
                        pyautogui.press('return')
                        time.sleep(1)
                    
                    # Cerrar la ventana de búsqueda
                    logger.info("Cerrando ventana de búsqueda...")
                    time.sleep(0.5)
                    self.close_expediente_window()
                    
                    logger.warning("Diálogo de error cerrado, retornando False")
                    return False
            except Exception:
                pass
            
            # Patrón 2: Buscar en descendientes del app_window
            try:
                descendants = app_window.descendants()
                for descendant in descendants:
                    try:
                        desc_text = descendant.window_text()
                        # Buscar el mensaje de error específico
                        if "no es válido" in desc_text.lower():
                            logger.warning(f"EXPEDIENTE INVÁLIDO! Mensaje: {desc_text}")
                            
                            # Actualizar Excel
                            logger.info("Registrando NRO EXPEDIENTE INVALIDO en Excel...")
                            self.update_excel_executor_result("NRO EXPEDIENTE INVALIDO")
                            
                            # Buscar botón "Aceptar"
                            try:
                                aceptar_btn = app_window.child_window(title="Aceptar")
                                if aceptar_btn.exists(timeout=2):
                                    logger.info("Botón 'Aceptar' encontrado, haciendo clic...")
                                    aceptar_btn.invoke()
                                    time.sleep(1)
                            except Exception as e:
                                logger.warning(f"No se encontró botón: {e}")
                                pyautogui.press('return')
                                time.sleep(1)
                            
                            # Cerrar ventana
                            logger.info("Cerrando ventana de búsqueda...")
                            time.sleep(0.5)
                            self.close_expediente_window()
                            
                            return False
                    except Exception:
                        pass
            except Exception:
                pass
            
            logger.info("No se encontró mensaje de error, expediente es VÁLIDO")
            return True
        
        except Exception as e:
            logger.warning(f"Error verificando expediente: {e}")
            logger.info("Asumiendo que el expediente es válido")
            return True
    
    def check_expediente_error_screen(self):
        """
        Verifica si hay mensaje de error usando MSAA.
        Si hay error, actualiza Excel, cierra y retorna False para reintentar.
        Si es válido, retorna True para proceder a validar ejecutor.
        """
        logger.info("Verificando si el expediente es válido...")
        
        try:
            # Buscar diálogo de error por MSAA
            logger.info("Buscando diálogo de error por MSAA...")
            desktop = Desktop(backend="uia")
            
            # Buscar diálogos que contengan "Error" en el título
            try:
                error_dialogs = []
                
                # Patrón 1: "Selección de Expediente Coactivo - Error"
                try:
                    win = desktop.window(title_re=".*Expediente.*Error.*")
                    if win.exists(timeout=1):
                        error_dialogs.append(win)
                        logger.warning("EXPEDIENTE INVÁLIDO! Diálogo de error detectado por MSAA")
                except Exception:
                    pass
                
                # Patrón 2: Solo "Error"
                if not error_dialogs:
                    try:
                        win = desktop.window(title_re=".*Error.*")
                        if win.exists(timeout=1):
                            error_dialogs.append(win)
                            logger.warning("EXPEDIENTE INVÁLIDO! Diálogo de error detectado por MSAA")
                    except Exception:
                        pass
                
                if error_dialogs:
                    # Actualizar Excel marcando como inválido
                    logger.info("Registrando NRO EXPEDIENTE INVALIDO en Excel...")
                    self.update_excel_executor_result("NRO EXPEDIENTE INVALIDO")
                    
                    # Presionar Enter para cerrar el diálogo
                    logger.info("Presionando Enter para cerrar el diálogo de error...")
                    pyautogui.press('return')
                    time.sleep(1)
                    
                    # Cerrar la ventana de búsqueda
                    logger.info("Cerrando ventana de búsqueda...")
                    time.sleep(0.5)
                    self.close_expediente_window()
                    
                    return False  # Expediente inválido
            
            except Exception as msaa_e:
                logger.warning(f"Error en búsqueda MSAA: {msaa_e}")
            
            # Si no se detectó error, asumir que es válido
            logger.info("No se detectó mensaje de error, expediente es VÁLIDO")
            return True
        
        except Exception as e:
            logger.warning(f"Error verificando pantalla: {e}")
            logger.info("Asumiendo que el expediente es válido")
            return True
    
    def close_expediente_window(self):
        """
        Cierra la ventana de búsqueda de expediente usando Escape.
        """
        try:
            logger.info("Intentando encontrar botón 'Cerrar' con MSAA...")
            desktop = Desktop(backend="uia")
            
            # Buscar ventana de expediente
            try:
                win = desktop.window(title_re=".*Expediente.*")
                if win.exists(timeout=1):
                    # Buscar botón Cerrar
                    try:
                        cerrar_btn = win.child_window(title_re="Cerrar|Cancelar", control_type="Button")
                        if cerrar_btn.exists(timeout=1):
                            logger.info("Botón 'Cerrar' encontrado con MSAA, haciendo clic...")
                            cerrar_btn.invoke()
                            time.sleep(1)
                            logger.info("Ventana cerrada exitosamente con MSAA")
                            return True
                    except Exception as btn_e:
                        logger.warning(f"No se encontró botón con MSAA: {btn_e}")
            except Exception as win_e:
                logger.warning(f"No se encontró ventana: {win_e}")
        except Exception as msaa_e:
            logger.warning(f"Error en búsqueda MSAA: {msaa_e}")
        
        # Último fallback: presionar Escape para cerrar
        logger.info("Fallback final: presionando Escape para cerrar ventana...")
        pyautogui.press('escape')
        time.sleep(1)
        logger.info("Ventana cerrada con Escape")
        return True
    
    def validate_executor(self):
        """
        Presiona ALT+A para continuar con el proceso de embargo.
        
        Nota: Si el tipo es IEI, la detección y extracción de RUC
        ocurre después de ALT+A en la función fill_interventor_and_plazo()
        """
        logger.info("Presionando ALT+A para continuar con el proceso de embargo...")
        
        try:
            time.sleep(0.5)
            pyautogui.hotkey('alt', 'a')
            logger.info(" ALT+A presionado correctamente")
            time.sleep(1)
            
            return True
        
        except Exception as e:
            logger.error(f"Error presionando ALT+A: {str(e)}")
            return False
    
    def click_proceso_embargo(self):
        """
        Busca y hace clic en 'Proceso de Embargo' en el menú.
        Utiliza MSAA para encontrar el elemento. El elemento está en el árbol de controles.
        Similar al patrón que funciona en click_trabar_embargo().
        """
        logger.info("Buscando 'Proceso de Embargo' en el menú...")
        
        try:
            desktop = Desktop(backend="uia")
            
            # Buscar ventana principal de SIRAT (Menú de Opciones)
            app = None
            try:
                app = desktop.window(title_re=".*Menú.*")
                if not app.exists(timeout=2):
                    app = None
            except:
                app = None
            
            if not app:
                try:
                    app = desktop.window(title_re=".*SIRAT.*")
                    if not app.exists(timeout=2):
                        app = None
                except:
                    app = None
            
            if not app:
                try:
                    app = desktop.active()
                except:
                    app = None
            
            if app:
                logger.info("Ventana encontrada, iterando descendientes...")
                
                try:
                    descendants = app.descendants()
                    logger.info(f"Total de descendientes: {len(descendants)}")
                    
                    # Listado de debug: mostrar elementos con "Proceso" o "Embargo"
                    elementos_embargo = []
                    
                    for idx, descendant in enumerate(descendants):
                        try:
                            # Obtener texto con strip() para eliminar espacios en blanco ocultos
                            desc_text = descendant.window_text().strip()
                            
                            # Buscar elementos que contengan "Embargo"
                            if "Embargo" in desc_text:
                                elementos_embargo.append((idx, desc_text))
                                logger.info(f"[{idx}] Elemento con 'Embargo': '{desc_text}'")
                                
                                # BÚSQUEDA EXACTA: "Proceso de Embargo"
                                if desc_text == "Proceso de Embargo":
                                    logger.info(f" 'Proceso de Embargo' encontrado (búsqueda exacta)")
                                    rect = descendant.rectangle()
                                    click_x = (rect.left + rect.right) // 2
                                    click_y = (rect.top + rect.bottom) // 2
                                    
                                    if click_x > 0 and click_y > 0:
                                        logger.info(f"Coordenadas válidas: ({click_x}, {click_y})")
                                        
                                        # Guardar coordenadas para reutilizar después
                                        self.proceso_embargo_coords = (click_x, click_y)
                                        logger.info(f" Coordenadas de 'Proceso de Embargo' guardadas: {self.proceso_embargo_coords}")
                                        
                                        pyautogui.click(click_x, click_y)
                                        logger.info("Esperando 2 segundos para que se expanda el menú...")
                                        time.sleep(2)
                                        logger.info(" Clic en Proceso de Embargo completado")
                                        return True
                        
                        except Exception as e:
                            pass
                    
                    # Si no encontró exacto, buscar por palabras clave
                    if not elementos_embargo:
                        logger.info("No se encontraron elementos con 'Embargo'")
                    else:
                        logger.info(f"Se encontraron {len(elementos_embargo)} elementos con 'Embargo', pero ninguno coincide exactamente")
                
                except Exception as e:
                    logger.warning(f"Error iterando descendientes: {e}")
            
            logger.warning("No se pudo encontrar 'Proceso de Embargo'")
            return False
        
        except Exception as e:
            logger.error(f"Error en click_proceso_embargo: {str(e)}")
            return False
    
    def click_trabar_embargo(self):
        """
        Busca y hace clic en 'Trabar Embargo' usando el patrón que funciona para otros elementos.
        El elemento tiene espacios en blanco al final que se deben eliminar con strip().
        """
        logger.info("Buscando 'Trabar Embargo' (usando patrón de descendientes)...")
        
        try:
            desktop = Desktop(backend="uia")
            
            # Buscar ventana SIRAT
            app = None
            try:
                app = desktop.window(title_re=".*SIRAT.*")
                if not app.exists(timeout=2):
                    app = None
            except:
                app = None
            
            if not app:
                try:
                    app = desktop.active()
                except:
                    app = None
            
            if app:
                logger.info("Ventana encontrada, iterando descendientes...")
                
                try:
                    descendants = app.descendants()
                    logger.info(f"Total de descendientes: {len(descendants)}")
                    
                    # Listas para debugging
                    elementos_embargo = []
                    elementos_trabar = []
                    
                    # Iterar todos los descendientes y buscar "Trabar Embargo"
                    for i, descendant in enumerate(descendants):
                        try:
                            desc_text = descendant.window_text().strip()  # ← IMPORTANTE: strip()
                            
                            # Recolectar elementos relevantes para debugging
                            if "embargo" in desc_text.lower():
                                elementos_embargo.append((i, desc_text))
                            
                            if "trabar" in desc_text.lower():
                                elementos_trabar.append((i, desc_text))
                            
                            # Búsqueda exacta con strip()
                            if desc_text == "Trabar Embargo":
                                logger.info(f" 'Trabar Embargo' encontrado (exacto) en índice {i}")
                                rect = descendant.rectangle()
                                
                                if rect.width() > 0 and rect.height() > 0:
                                    click_x = (rect.left + rect.right) // 2
                                    click_y = (rect.top + rect.bottom) // 2
                                    logger.info(f"Coordenadas válidas: ({click_x}, {click_y})")
                                    
                                    # Guardar coordenadas para reutilizar después
                                    self.trabar_embargo_coords = (click_x, click_y)
                                    logger.info(f" Coordenadas de 'Trabar Embargo' guardadas: {self.trabar_embargo_coords}")
                                    
                                    logger.info(f"Haciendo clic en: ({click_x}, {click_y})")
                                    pyautogui.click(click_x, click_y)
                                    time.sleep(1)
                                    logger.info(" Clic completado exitosamente")
                                    return True
                                else:
                                    # Intentar invoke si no tiene coordenadas válidas
                                    logger.info("Coordenadas inválidas, intentando invoke...")
                                    try:
                                        descendant.invoke()
                                        time.sleep(1)
                                        logger.info(" Invocado correctamente")
                                        return True
                                    except:
                                        pyautogui.press('return')
                                        time.sleep(1)
                                        logger.info(" Enter presionado")
                                        return True
                        
                        except Exception as e:
                            pass
                    
                    # Si no se encontró, mostrar debugging detallado
                    logger.warning("'Trabar Embargo' NO encontrado en descendientes")
                    logger.warning("=" * 70)
                    logger.warning("DEBUG: ELEMENTOS CON 'EMBARGO':")
                    logger.warning("=" * 70)
                    for idx, elem_text in elementos_embargo:
                        logger.warning(f"[{idx}] '{elem_text}'")
                    
                    if not elementos_embargo:
                        logger.warning("(Ninguno encontrado)")
                    
                    logger.warning("=" * 70)
                    logger.warning("DEBUG: ELEMENTOS CON 'TRABAR':")
                    logger.warning("=" * 70)
                    for idx, elem_text in elementos_trabar:
                        logger.warning(f"[{idx}] '{elem_text}'")
                    
                    if not elementos_trabar:
                        logger.warning("(Ninguno encontrado)")
                    
                    logger.warning("=" * 70)
                    logger.warning("Esto sugiere que la estructura del menú es diferente en DSE vs IEI")
                    logger.warning("=" * 70)
                
                except Exception as e:
                    logger.warning(f"Error iterando: {e}")
            
            logger.warning("No se pudo encontrar 'Trabar Embargo'")
            return False
        
        except Exception as e:
            logger.error(f"Error en click_trabar_embargo: {str(e)}")
            return False
    
    def click_trabar_intervencion_informacion(self):
        """
        Busca y hace DOBLE CLIC en 'Trabar Intervención en Información' usando coordenadas.
        Similar al patrón que funciona en click_trabar_embargo().
        El elemento tiene espacios en blanco al final que deben eliminarse con strip().
        """
        logger.info("Buscando 'Trabar Intervención en Información' en el menú...")
        
        try:
            desktop = Desktop(backend="uia")
            
            # Buscar ventana SIRAT
            app = None
            try:
                app = desktop.window(title_re=".*SIRAT.*")
                if not app.exists(timeout=2):
                    app = None
            except:
                app = None
            
            if not app:
                try:
                    app = desktop.active()
                except:
                    app = None
            
            if app:
                logger.info("Ventana encontrada, iterando descendientes...")
                
                try:
                    descendants = app.descendants()
                    logger.info(f"Total de descendientes: {len(descendants)}")
                    
                    # Iterar todos los descendientes y buscar "Trabar Intervención en Información"
                    for i, descendant in enumerate(descendants):
                        try:
                            desc_text = descendant.window_text().strip()  # ← IMPORTANTE: strip()
                            
                            # Log detallado de elementos que contienen "Intervención"
                            if "intervención" in desc_text.lower():
                                logger.info(f"[{i}] Elemento con 'Intervención': '{desc_text}'")
                            
                            # Búsqueda exacta con strip()
                            if desc_text == "Trabar Intervención en Información":
                                logger.info(f" 'Trabar Intervención en Información' encontrado (exacto) en índice {i}")
                                rect = descendant.rectangle()
                                
                                if rect.width() > 0 and rect.height() > 0:
                                    click_x = (rect.left + rect.right) // 2
                                    click_y = (rect.top + rect.bottom) // 2
                                    logger.info(f"Coordenadas válidas: ({click_x}, {click_y})")
                                    logger.info(f"Haciendo DOBLE CLIC en: ({click_x}, {click_y})")
                                    pyautogui.doubleClick(click_x, click_y)
                                    time.sleep(1.5)
                                    logger.info(" Doble clic completado exitosamente")
                                    return True
                                else:
                                    # Intentar invoke si no tiene coordenadas válidas
                                    logger.info("Coordenadas inválidas, intentando invoke...")
                                    try:
                                        descendant.invoke()
                                        time.sleep(1)
                                        logger.info(" Invocado correctamente")
                                        return True
                                    except:
                                        pyautogui.press('return')
                                        time.sleep(1)
                                        logger.info(" Enter presionado")
                                        return True
                        
                        except Exception as e:
                            pass
                    
                    logger.warning("'Trabar Intervención en Información' no encontrado en descendientes")
                    logger.info("Listando TODOS los elementos con 'trabar' para debug...")
                    for i, descendant in enumerate(descendants):
                        try:
                            desc_text = descendant.window_text().strip()
                            if "trabar" in desc_text.lower():
                                logger.info(f"  [{i}] '{desc_text}'")
                        except:
                            pass
                
                except Exception as e:
                    logger.warning(f"Error iterando: {e}")
            
            logger.warning("No se pudo encontrar 'Trabar Intervención en Información'")
            return False
        
        except Exception as e:
            logger.error(f"Error en click_trabar_intervencion_informacion: {str(e)}")
            return False
    
    def handle_trabar_intervencion_aviso(self):
        """
        Presiona ENTER directamente después del doble clic en 'Trabar Intervención en Información'.
        
        Flujo:
        1. Esperar 1 segundo de timeout
        2. Presionar ENTER
        3. Esperar 1 segundo adicional
        4. Retornar True para continuar con INTERVENTOR y PLAZO
        5. ALT+S se presionará después de completar ambos campos
        """
        logger.info("\nPresionando ENTER después de 'Trabar Intervención en Información'...")
        
        try:
            # Timeout de 1 segundo antes de presionar ENTER
            logger.info("Esperando 1 segundo...")
            time.sleep(1)
            
            logger.info("Presionando ENTER...")
            pyautogui.press('return')
            
            # Esperar 1 segundo después de ENTER
            logger.info("Esperando 1 segundo después de ENTER...")
            time.sleep(1)
            
            logger.info(" ENTER presionado correctamente")
            logger.info("Continuando con el flujo de INTERVENTOR y PLAZO...")
            
            return True
        
        except Exception as e:
            logger.error(f"Error en handle_trabar_intervencion_aviso: {str(e)}")
            # No fallar si hay error - continuar de todas formas
            return True
    
    def handle_trabar_deposito_aviso(self):
        """
        Presiona ENTER directamente después del doble clic en 'Trabar Depósito sin Extracción'.
        Idéntico a handle_trabar_intervencion_aviso() para mantener consistencia.
        
        Flujo:
        1. Esperar 1 segundo de timeout
        2. Presionar ENTER
        3. Esperar 1 segundo adicional
        4. Retornar True para continuar con MONTO
        """
        logger.info("\nPresionando ENTER después de 'Trabar Depósito sin Extracción'...")
        
        try:
            # Timeout de 1 segundo antes de presionar ENTER
            logger.info("Esperando 1 segundo...")
            time.sleep(1)
            
            logger.info("Presionando ENTER...")
            pyautogui.press('return')
            
            # Esperar 1 segundo después de ENTER
            logger.info("Esperando 1 segundo después de ENTER...")
            time.sleep(1)
            
            logger.info(" ENTER presionado correctamente")
            logger.info("Continuando con el flujo de MONTO...")
            
            return True
        
        except Exception as e:
            logger.error(f"Error en handle_trabar_deposito_aviso: {str(e)}")
            # No fallar si hay error - continuar de todas formas
            return True
    
    def click_trabar_deposito_sin_extraccion(self):
        """
        Busca y hace DOBLE CLIC en 'Trabar Depósito sin Extracción' usando coordenadas.
        Similar al patrón que funciona en click_trabar_embargo().
        Se usa cuando la dependencia es 23 (MEPECO).
        El elemento tiene espacios en blanco al final que deben eliminarse con strip().
        """
        logger.info("Buscando 'Trabar Depósito sin Extracción' en el menú...")
        
        try:
            desktop = Desktop(backend="uia")
            
            # Buscar ventana SIRAT
            app = None
            try:
                app = desktop.window(title_re=".*SIRAT.*")
                if not app.exists(timeout=2):
                    app = None
            except:
                app = None
            
            if not app:
                try:
                    app = desktop.active()
                except:
                    app = None
            
            if app:
                logger.info("Ventana encontrada, iterando descendientes...")
                
                try:
                    descendants = app.descendants()
                    logger.info(f"Total de descendientes: {len(descendants)}")
                    
                    # Iterar todos los descendientes y buscar "Trabar Depósito sin Extracción"
                    for i, descendant in enumerate(descendants):
                        try:
                            desc_text = descendant.window_text().strip()  # ← IMPORTANTE: strip()
                            
                            # Log detallado de elementos que contienen "Depósito"
                            if "depósito" in desc_text.lower():
                                logger.info(f"[{i}] Elemento con 'Depósito': '{desc_text}'")
                            
                            # Búsqueda exacta con strip()
                            if desc_text == "Trabar Depósito sin Extracción":
                                logger.info(f" 'Trabar Depósito sin Extracción' encontrado (exacto) en índice {i}")
                                rect = descendant.rectangle()
                                
                                if rect.width() > 0 and rect.height() > 0:
                                    click_x = (rect.left + rect.right) // 2
                                    click_y = (rect.top + rect.bottom) // 2
                                    logger.info(f"Coordenadas válidas: ({click_x}, {click_y})")
                                    logger.info(f"Haciendo DOBLE CLIC en: ({click_x}, {click_y})")
                                    pyautogui.doubleClick(click_x, click_y)
                                    time.sleep(1.5)
                                    logger.info(" Doble clic completado exitosamente")
                                    return True
                                else:
                                    # Intentar invoke si no tiene coordenadas válidas
                                    logger.info("Coordenadas inválidas, intentando invoke...")
                                    try:
                                        descendant.invoke()
                                        time.sleep(1)
                                        logger.info(" Invocado correctamente")
                                        return True
                                    except:
                                        pyautogui.press('return')
                                        time.sleep(1)
                                        logger.info(" Enter presionado")
                                        return True
                        
                        except Exception as e:
                            pass
                    
                    logger.warning("'Trabar Depósito sin Extracción' no encontrado en descendientes")
                    logger.info("Listando TODOS los elementos con 'trabar' para debug...")
                    for i, descendant in enumerate(descendants):
                        try:
                            desc_text = descendant.window_text().strip()
                            if "trabar" in desc_text.lower():
                                logger.info(f"  [{i}] '{desc_text}'")
                        except:
                            pass
                
                except Exception as e:
                    logger.warning(f"Error iterando: {e}")
            
            logger.warning("No se pudo encontrar 'Trabar Depósito sin Extracción'")
            return False
        
        except Exception as e:
            logger.error(f"Error en click_trabar_deposito_sin_extraccion: {str(e)}")
            return False
    
    def handle_post_embargo_flow(self):
        """
        Maneja el flujo después de 'Trabar Embargo' según el TIPO DE MEDIDA del primer expediente.
        
        IMPORTANTE: El tipo (IEI o DSE) se determina leyendo la columna TIPO DE MEDIDA del Excel,
        NO según el número de dependencia (21 o 23). Ambas dependencias pueden contener ambos tipos.
        
        - TIPO DE MEDIDA = IEI: Ejecuta 'Trabar Intervención en Información'Rellena INTERVENTOR y PLAZO
        - TIPO DE MEDIDA = DSE: Ejecuta 'Trabar Depósito sin Extracción'Rellena MONTO
        """
        logger.info("\n" + "=" * 70)
        logger.info("MANEJANDO FLUJO POST-EMBARGO (PRIMER EXPEDIENTE)")
        logger.info("=" * 70)
        
        try:
            # ================================================================
            # PASO 1: DETECTAR TIPO LEYENDO TIPO DE MEDIDA DEL PRIMER EXPEDIENTE
            # ================================================================
            logger.info("Detectando tipo del primer expediente...")
            
            excel_file = SCRIPT_DIR / "EXPEDIENTES.xlsx"
            expedientes = pd.read_excel(excel_file, engine="openpyxl", dtype=str)
            
            if "TIPO DE MEDIDA" not in expedientes.columns:
                logger.error("La columna 'TIPO DE MEDIDA' no existe en el Excel")
                return False
            
            tipo_medida = str(expedientes.iloc[0]["TIPO DE MEDIDA"]).strip().upper()
            logger.info(f"Tipo de Medida del primer expediente: {tipo_medida}")
            
            # Determinar si es IEI o DSE
            if "IEI" in tipo_medida:
                tipo_detectado = "IEI"
            elif "DSE" in tipo_medida:
                tipo_detectado = "DSE"
            else:
                logger.error(f"Tipo de medida no reconocido: {tipo_medida}")
                return False
            
            logger.info(f"Tipo detectado: {tipo_detectado}")
            logger.info(f"Dependencia: {self.dependencia}")
            
            # ================================================================
            # PASO 2: EJECUTAR FLUJO SEGÚN TIPO DETECTADO
            # ================================================================
            if tipo_detectado == "IEI":
                logger.info(" Ejecutando flujo IEI: Trabar Intervención en Información")
                
                # Hacer clic en Trabar Intervención en Información
                if not self.click_trabar_intervencion_informacion():
                    logger.error("No se pudo hacer clic en Trabar Intervención en Información")
                    return False
                
                # Manejar el aviso que puede aparecer
                logger.info("\nManejando posible aviso...")
                if not self.handle_trabar_intervencion_aviso():
                    logger.error("Error manejando el aviso")
                    return False
                
                # Rellenar campos INTERVENTOR y PLAZO
                logger.info("\nRellenando campos INTERVENTOR y PLAZO...")
                return self.fill_interventor_and_plazo()
            
            elif tipo_detectado == "DSE":
                logger.info(" Ejecutando flujo DSE: Trabar Depósito sin Extracción")
                
                # Hacer clic en Trabar Depósito sin Extracción
                if not self.click_trabar_deposito_sin_extraccion():
                    logger.error("No se pudo hacer clic en Trabar Depósito sin Extracción")
                    return False
                
                # Manejar el aviso que puede aparecer
                logger.info("\nManejando posible aviso...")
                if not self.handle_trabar_deposito_aviso():
                    logger.error("Error manejando el aviso")
                    return False
                
                # Rellenar campo MONTO
                logger.info("\nRellenando campo MONTO...")
                return self.fill_monto()
        
        except Exception as e:
            logger.error(f"Error en handle_post_embargo_flow: {str(e)}")
            return False
    
    def fill_interventor_and_plazo(self):
        """
        Llena los campos de INTERVENTOR y PLAZO en el formulario secuencialmente.
        
        Flujo:
        1. Busca y selecciona el campo INTERVENTOR (por MSAA o imagen)
        2. Hace clic en el campo INTERVENTOR
        3. Digita el código de INTERVENTOR desde Excel
        4. Presiona Enter
        5. Busca y selecciona el campo PLAZO
        6. Hace clic en el campo PLAZO
        7. Digita el PLAZO desde Excel
        8. Presiona Enter
        
        Las capturas de referencia son "1. Interventor.png" y "2. Plazo.png"
        """
        logger.info("Rellenando campos de INTERVENTOR y PLAZO...")
        
        try:
            # Cargar datos del Excel
            excel_file = SCRIPT_DIR / "EXPEDIENTES.xlsx"
            expedientes = pd.read_excel(excel_file, engine="openpyxl", dtype=str)
            
            # Obtener valores de INTERVENTOR y PLAZO del primer expediente válido
            interventor = None
            plazo = None
            
            if "INTERVENTOR" in expedientes.columns:
                interventor = str(expedientes.iloc[self.primer_expediente_idx]["INTERVENTOR"]).strip()
                logger.info(f"INTERVENTOR obtenido del Excel (fila {self.primer_expediente_idx + 2}): '{interventor}'")
            else:
                logger.warning("No existe columna 'INTERVENTOR' en el Excel")
                interventor = ""
            
            if "PLAZO" in expedientes.columns:
                plazo = str(expedientes.iloc[self.primer_expediente_idx]["PLAZO"]).strip()
                logger.info(f"PLAZO obtenido del Excel (fila {self.primer_expediente_idx + 2}): '{plazo}'")
            else:
                logger.warning("No existe columna 'PLAZO' en el Excel")
                plazo = ""
            
            # ============================================================
            # PASO 1: SELECCIONAR Y HACER CLIC EN CAMPO INTERVENTOR
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 1: Esperando 0.5s y digitando INTERVENTOR")
            logger.info("=" * 70)
            
            # Esperar 0.5 segundos antes de digitar INTERVENTOR
            logger.info("Esperando 0.5 segundos antes de digitar INTERVENTOR...")
            time.sleep(0.5)
            
            # Digitar INTERVENTOR directamente (sin hacer clic en el campo)
            logger.info(f"Digitando INTERVENTOR: '{interventor}'")
            pyautogui.write(interventor, interval=0.05)
            time.sleep(0.2)
            
            # ============================================================
            # PASO 2: Presionar TAB para pasar a PLAZO
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 2: Presionando TAB para pasar a PLAZO")
            logger.info("=" * 70)
            
            logger.info("Presionando TAB para pasar al campo PLAZO...")
            pyautogui.press('tab')
            time.sleep(0.2)
            
            # ============================================================
            # PASO 3: Digitar PLAZO
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 3: Digitando PLAZO")
            logger.info("=" * 70)
            
            logger.info(f"Digitando PLAZO: '{plazo}'")
            pyautogui.write(plazo, interval=0.05)
            time.sleep(0.3)
            
            # ============================================================
            # PASO 4: Verificar valores via MSAA
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 4: Verificando que los valores aparecieron en los campos via MSAA")
            logger.info("=" * 70)
            
            try:
                desktop = Desktop(backend="uia")
                
                # Buscar ventana principal de SIRAT
                sirat_window = None
                try:
                    sirat_windows = desktop.windows(title_re=".*SIRAT.*")
                    if sirat_windows:
                        sirat_window = sirat_windows[0]
                except:
                    # Si no encontramos por título, usar ventana activa
                    sirat_window = desktop.active()
                
                if sirat_window and sirat_window.exists(timeout=1):
                    # Buscar campo INTERVENTOR
                    try:
                        interventor_field = sirat_window.child_window(
                            title="INTERVENTOR",
                            control_type="Edit"
                        )
                        if interventor_field.exists(timeout=1):
                            valor_interventor = interventor_field.window_text()
                            logger.info(f"Valor en campo INTERVENTOR: '{valor_interventor}'")
                            
                            if interventor in valor_interventor:
                                logger.info(" INTERVENTOR se ingresó correctamente")
                            else:
                                logger.warning(f"INTERVENTOR ingresado no coincide: esperado '{interventor}', obtenido '{valor_interventor}'")
                    except Exception as e:
                        logger.info(f"No se pudo verificar INTERVENTOR via MSAA: {e}")
                    
                    # Buscar campo PLAZO
                    try:
                        plazo_field = sirat_window.child_window(
                            title="PLAZO",
                            control_type="Edit"
                        )
                        if plazo_field.exists(timeout=1):
                            valor_plazo = plazo_field.window_text()
                            logger.info(f"Valor en campo PLAZO: '{valor_plazo}'")
                            
                            if plazo in valor_plazo:
                                logger.info(" PLAZO se ingresó correctamente")
                            else:
                                logger.warning(f"PLAZO ingresado no coincide: esperado '{plazo}', obtenido '{valor_plazo}'")
                    except Exception as e:
                        logger.info(f"No se pudo verificar PLAZO via MSAA: {e}")
            except Exception as e:
                logger.info(f"No se pudo verificar valores via MSAA: {e}")
            
            logger.info("=" * 70)
            logger.info(" CAMPOS INTERVENTOR Y PLAZO COMPLETADOS EXITOSAMENTE")
            logger.info("=" * 70)
            
            # ============================================================
            # PASO 5: Presionar ALT+A después de completar ambos campos
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 5: Presionando ALT+A para confirmar INTERVENTOR y PLAZO")
            logger.info("=" * 70)
            
            logger.info("Presionando ALT+A...")
            pyautogui.hotkey('alt', 'a')
            time.sleep(1)
            logger.info(" ALT+A presionado correctamente")
            
            # ============================================================
            # PASO 6: Detectar mensaje "El Expediente" y presionar ALT+S si existe
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 6: Verificando si hay mensaje de aviso 'El Expediente'")
            logger.info("=" * 70)
            
            aviso_detected, aviso_mensaje = self.detect_expediente_aviso(timeout=2)
            
            if aviso_detected:
                logger.warning(" MENSAJE DE AVISO 'EL EXPEDIENTE' DETECTADO")
                logger.warning(f"Contenido: {aviso_mensaje[:100]}...")
            else:
                logger.info(" No hay mensaje de aviso con 'El Expediente' (posible mensaje: ¿Desea Ud. grabar la Resolución?)")
            
            # NOTA IMPORTANTE: SIEMPRE aparecerá un mensaje-aviso en este punto:
            # 1. Si hay embargos activos: "El Expediente..." (detectado arriba)
            # 2. Si NO hay embargos: "¿Desea Ud. grabar la Resolución?" (no detectado por el patrón anterior, pero presente)
            # En ambos casos se presiona ALT+S para continuar
            logger.info("Presionando ALT+S para cerrar el aviso (ya sea 'El Expediente' o '¿Desea Ud. grabar la Resolución?')...")
            time.sleep(0.5)
            pyautogui.hotkey('alt', 's')
            time.sleep(1)
            logger.info(" ALT+S presionado correctamente (1er ALT+S)")
            
            # ============================================================
            # PASO 7: Pequeño timeout antes del siguiente ALT+S
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 7: Esperando timeout antes del siguiente ALT+S")
            logger.info("=" * 70)
            
            logger.info("Esperando 0.5 segundos...")
            time.sleep(0.5)
            
            # ============================================================
            # PASO 8: Presionar ALT+S nuevamente para continuar
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 8: Presionando ALT+S para continuar al siguiente paso (2do ALT+S)")
            logger.info("=" * 70)
            
            logger.info("Presionando ALT+S para continuar...")
            pyautogui.hotkey('alt', 's')
            time.sleep(1)
            logger.info(" ALT+S presionado correctamente (2do ALT+S)")
            
            # ============================================================
            # PASO 9: Detectar mensaje de Resolución Coactiva
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 9: Verificando si aparece mensaje de Resolución Coactiva")
            logger.info("=" * 70)
            
            rc_detected, rc_mensaje = self.detect_resolucion_coactiva_aviso(timeout=2)
            
            if rc_detected:
                logger.warning(" MENSAJE DE RESOLUCIÓN COACTIVA DETECTADO")
                logger.warning(f"Contenido: {rc_mensaje[:100]}...")
                
                # Extraer el número de Resolución Coactiva
                logger.info("Extrayendo número de Resolución Coactiva...")
                rc_number = self.extract_resolucion_coactiva_number(rc_mensaje)
                
                if rc_number:
                    logger.info(f" Número de RC EXTRAÍDO EXITOSAMENTE: {rc_number}")
                    
                    # Guardar el número de RC en Excel como TEXTO (preservando el 0 inicial)
                    logger.info(f"Guardando número de RC en columna RESULTADO del Excel como texto...")
                    # Usar openpyxl para guardar como texto sin la comilla
                    self.update_excel_result(rc_number)
                    logger.info(f" Número de RC guardado en Excel: {rc_number}")
                else:
                    logger.warning("✗ No se pudo extraer número de RC del mensaje")
                    self.update_excel_result("ERROR: No se extrajo número de RC")
            else:
                logger.warning("✗ No se detectó mensaje de Resolución Coactiva")
                self.update_excel_result("ERROR: No se detectó RC")
            
            # ============================================================
            # PASO 10: Presionar ENTER para aceptar
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 10: Presionando ENTER para aceptar")
            logger.info("=" * 70)
            
            logger.info("Presionando ENTER...")
            pyautogui.press('return')
            time.sleep(1)
            logger.info(" ENTER presionado correctamente")
            
            # ============================================================
            # PASO 11: Presionar ALT+C para regresar al menú
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 11: Presionando ALT+C para regresar al menú")
            logger.info("=" * 70)
            
            logger.info("Presionando ALT+C para regresar al menú...")
            time.sleep(0.5)
            pyautogui.hotkey('alt', 'c')
            time.sleep(1)
            logger.info(" ALT+C presionado correctamente")
            
            # ============================================================
            # PASO 12: Eliminar desplazamientos del menú para acceder a "Accesos"
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 12: Eliminando desplazamientos del menú")
            logger.info("=" * 70)
            
            # Clic en "Trabar Embargo" para eliminar un desplazamiento
            logger.info("Haciendo clic en 'Trabar Embargo' para eliminar desplazamiento...")
            if not self.click_trabar_embargo():
                logger.warning("No se pudo hacer clic en 'Trabar Embargo' (continuando...)")
            time.sleep(0.5)
            
            # Clic en "Proceso de Embargo" para eliminar otro desplazamiento
            logger.info("Haciendo clic en 'Proceso de Embargo' para eliminar desplazamiento...")
            if not self.click_proceso_embargo():
                logger.warning("No se pudo hacer clic en 'Proceso de Embargo' (continuando...)")
            time.sleep(0.5)
            
            logger.info(" Desplazamientos eliminados - 'Accesos' ahora visible en el mismo nivel")
            
            # ============================================================
            # PASO 13: Hacer clic en "Accesos"
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 13: Haciendo clic en 'Accesos'")
            logger.info("=" * 70)
            
            # Buscar y hacer clic en "Accesos" directamente (sin desplazamiento)
            logger.info("Buscando 'Accesos' en el menú...")
            if not self._click_accesos_direct():
                logger.error("No se pudo hacer clic en 'Accesos'")
                return False
            
            time.sleep(0.5)
            
            # ============================================================
            # PASO 14: Hacer doble clic en "Cambio de Expediente"
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 14: Haciendo doble clic en 'Cambio de Expediente'")
            logger.info("=" * 70)
            
            if not self.click_cambio_expediente():
                logger.error("No se pudo hacer clic en 'Cambio de Expediente'")
                return False
            
            time.sleep(0.5)
            
            # ============================================================
            # PASO 15: Iniciar bucle de expedientes
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 15: Iniciando bucle de búsqueda de expedientes restantes")
            logger.info("=" * 70)
            
            if not self.expediente_loop_iei():
                logger.error("Error en bucle de expedientes IEI")
                return False
            
            logger.info(" PROCESO COMPLETADO EXITOSAMENTE")
            
            return True
        
        except Exception as e:
            logger.error(f"Error en fill_interventor_and_plazo: {str(e)}")
            return False
    
    def desplazar_menu_para_accesos(self):
        """
        Desplaza el menú para que "Accesos" sea visible.
        Reutiliza las coordenadas guardadas de "Trabar Embargo" y "Proceso de Embargo".
        
        Flujo:
        1. Hacer clic en "Trabar Embargo" (usa coordenadas guardadas)
        2. Hacer clic en "Proceso de Embargo" (usa coordenadas guardadas)
        3. Esto desplazará el menú y hará que "Accesos" sea visible
        """
        logger.info("\n" + "=" * 70)
        logger.info("DESPLAZANDO MENÚ PARA VISUALIZAR 'ACCESOS'")
        logger.info("=" * 70)
        
        try:
            # PASO 1: Hacer clic en "Trabar Embargo" usando coordenadas guardadas
            if self.trabar_embargo_coords:
                logger.info(f"Paso 1: Haciendo clic en 'Trabar Embargo' (coordenadas: {self.trabar_embargo_coords})")
                pyautogui.click(self.trabar_embargo_coords[0], self.trabar_embargo_coords[1])
                time.sleep(1)
                logger.info(" Clic en 'Trabar Embargo' completado")
            else:
                logger.warning("Coordenadas de 'Trabar Embargo' no disponibles, buscando nuevamente...")
                if not self.click_trabar_embargo():
                    logger.error("No se pudo encontrar 'Trabar Embargo'")
                    return False
            
            # PASO 2: Hacer clic en "Proceso de Embargo" usando coordenadas guardadas
            if self.proceso_embargo_coords:
                logger.info(f"Paso 2: Haciendo clic en 'Proceso de Embargo' (coordenadas: {self.proceso_embargo_coords})")
                pyautogui.click(self.proceso_embargo_coords[0], self.proceso_embargo_coords[1])
                time.sleep(1)
                logger.info(" Clic en 'Proceso de Embargo' completado")
            else:
                logger.warning("Coordenadas de 'Proceso de Embargo' no disponibles, buscando nuevamente...")
                if not self.click_proceso_embargo():
                    logger.error("No se pudo encontrar 'Proceso de Embargo'")
                    return False
            
            logger.info(" Menú desplazado exitosamente")
            logger.info("=" * 70)
            return True
        
        except Exception as e:
            logger.error(f"Error en desplazar_menu_para_accesos: {str(e)}")
            return False
    
    def click_accesos(self):
        """
        Busca y hace clic en 'Accesos' usando MSAA (patrón robusto igual a click_trabar_embargo).
        El elemento 'Accesos' está en el mismo nivel que 'Cobranza Coactiva' en el menú.
        Nota: El nombre 'Accesos' tiene espacios en blanco al final que se eliminan con .strip().
        
        IMPORTANTE: Antes de buscar, desplaza el menú para que 'Accesos' sea visible.
        Este clic iniciará el bucle de búsqueda de expedientes para IEI (Dependencia 21).
        """
        logger.info("\n" + "=" * 70)
        logger.info("BUSCANDO 'ACCESOS' EN EL MENÚ")
        logger.info("=" * 70)
        
        try:
            # PASO 1: Desplazar el menú para que 'Accesos' sea visible
            logger.info("Paso 1: Desplazando menú para visualizar 'Accesos'...")
            if not self.desplazar_menu_para_accesos():
                logger.warning("No se pudo desplazar el menú, intentando búsqueda directa...")
            
            time.sleep(1)
            
            # PASO 2: Buscar 'Accesos' después del desplazamiento
            logger.info("Paso 2: Buscando 'Accesos' después del desplazamiento...")
            
            desktop = Desktop(backend="uia")
            
            # Buscar ventana SIRAT
            app = None
            try:
                app = desktop.window(title_re=".*SIRAT.*", class_name="TApplication")
            except:
                pass
            
            if not app:
                try:
                    app = desktop.window(title_re=".*Menú.*", class_name="TApplication")
                except:
                    pass
            
            if app:
                logger.info("Ventana SIRAT encontrada, buscando 'Accesos'...")
                
                try:
                    # Buscar descendientes del árbol de controles (patrón probado)
                    descendants = app.descendants()
                    logger.info(f"Total de descendientes: {len(descendants)}")
                    
                    for i, descendant in enumerate(descendants):
                        try:
                            desc_text = descendant.window_text().strip()  # ← IMPORTANTE: strip()
                            
                            # Búsqueda exacta con strip()
                            if desc_text == "Accesos":
                                logger.info(f" 'Accesos' encontrado (exacto) en índice {i}")
                                rect = descendant.rectangle()
                                
                                if rect.width() > 0 and rect.height() > 0:
                                    click_x = (rect.left + rect.right) // 2
                                    click_y = (rect.top + rect.bottom) // 2
                                    logger.info(f"Coordenadas válidas: ({click_x}, {click_y})")
                                    
                                    logger.info(f"Haciendo clic en: ({click_x}, {click_y})")
                                    pyautogui.click(click_x, click_y)
                                    time.sleep(1)
                                    logger.info(" Clic completado exitosamente")
                                    return True
                                else:
                                    # Intentar invoke si no tiene coordenadas válidas
                                    logger.info("Coordenadas inválidas, intentando invoke...")
                                    try:
                                        descendant.invoke()
                                        time.sleep(1)
                                        logger.info(" Invocado correctamente")
                                        return True
                                    except:
                                        pyautogui.press('return')
                                        time.sleep(1)
                                        logger.info(" Enter presionado")
                                        return True
                        
                        except Exception as e:
                            pass
                    
                    logger.warning("'Accesos' NO encontrado en descendientes")
                
                except Exception as e:
                    logger.error(f"Error iterando descendientes: {e}")
            
            logger.warning("No se pudo encontrar 'Accesos'")
            return False
        
        except Exception as e:
            logger.error(f"Error en click_accesos: {str(e)}")
            return False
    
    def click_cambio_expediente(self):
        """
        Busca y hace DOBLE CLIC en 'Cambio de Expediente' usando MSAA (patrón robusto igual a click_trabar_embargo).
        El elemento 'Cambio de Expediente' está dentro de 'Accesos' en el menú.
        Nota: El nombre puede tener espacios en blanco al final que se eliminan con .strip().
        Este doble clic es el primer paso del bucle de expedientes en IEI.
        """
        logger.info("Buscando 'Cambio de Expediente' en el menú...")
        
        try:
            desktop = Desktop(backend="uia")
            
            # Buscar ventana SIRAT
            app = None
            try:
                app = desktop.window(title_re=".*SIRAT.*", class_name="TApplication")
            except:
                pass
            
            if not app:
                try:
                    app = desktop.window(title_re=".*Menú.*", class_name="TApplication")
                except:
                    pass
            
            if app:
                logger.info("Ventana SIRAT encontrada, buscando 'Cambio de Expediente'...")
                
                try:
                    # Buscar descendientes del árbol de controles (patrón probado)
                    descendants = app.descendants()
                    logger.info(f"Total de descendientes: {len(descendants)}")
                    
                    for i, descendant in enumerate(descendants):
                        try:
                            desc_text = descendant.window_text().strip()  # ← IMPORTANTE: strip()
                            
                            # Búsqueda exacta con strip()
                            if desc_text == "Cambio de Expediente":
                                logger.info(f" 'Cambio de Expediente' encontrado (exacto) en índice {i}")
                                rect = descendant.rectangle()
                                
                                if rect.width() > 0 and rect.height() > 0:
                                    click_x = (rect.left + rect.right) // 2
                                    click_y = (rect.top + rect.bottom) // 2
                                    logger.info(f"Coordenadas válidas: ({click_x}, {click_y})")
                                    
                                    logger.info(f"Haciendo DOBLE CLIC en: ({click_x}, {click_y})")
                                    pyautogui.doubleClick(click_x, click_y)
                                    time.sleep(1)
                                    logger.info(" Doble clic completado exitosamente")
                                    return True
                                else:
                                    # Intentar invoke si no tiene coordenadas válidas
                                    logger.info("Coordenadas inválidas, intentando invoke...")
                                    try:
                                        descendant.invoke()
                                        time.sleep(1)
                                        logger.info(" Invocado correctamente")
                                        return True
                                    except:
                                        pyautogui.press('return')
                                        time.sleep(1)
                                        logger.info(" Enter presionado")
                                        return True
                        
                        except Exception as e:
                            pass
                    
                    logger.warning("'Cambio de Expediente' NO encontrado en descendientes")
                
                except Exception as e:
                    logger.error(f"Error iterando descendientes: {e}")
            
            logger.warning("No se pudo encontrar 'Cambio de Expediente'")
            return False
        
        except Exception as e:
            logger.error(f"Error en click_cambio_expediente: {str(e)}")
            return False
    
    def _click_accesos_direct(self):
        """
        Busca y hace clic en 'Accesos' SIN desplazamiento previo.
        Se usa después de eliminar los desplazamientos del menú.
        Similar a click_accesos pero sin el paso de desplazamiento.
        """
        logger.info("Buscando 'Accesos' (sin desplazamiento previo)...")
        
        try:
            desktop = Desktop(backend="uia")
            
            # Buscar ventana SIRAT
            app = None
            try:
                app = desktop.window(title_re=".*SIRAT.*", class_name="TApplication")
            except:
                pass
            
            if not app:
                try:
                    app = desktop.window(title_re=".*Menú.*", class_name="TApplication")
                except:
                    pass
            
            if app:
                logger.info("Ventana SIRAT encontrada, buscando 'Accesos'...")
                
                try:
                    descendants = app.descendants()
                    
                    for descendant in descendants:
                        try:
                            desc_text = descendant.window_text().strip()
                            
                            if desc_text == "Accesos":
                                logger.info(f" 'Accesos' encontrado")
                                rect = descendant.rectangle()
                                
                                if rect.width() > 0 and rect.height() > 0:
                                    click_x = (rect.left + rect.right) // 2
                                    click_y = (rect.top + rect.bottom) // 2
                                    logger.info(f"Haciendo clic en 'Accesos': ({click_x}, {click_y})")
                                    pyautogui.click(click_x, click_y)
                                    time.sleep(1)
                                    logger.info(" Clic en 'Accesos' completado")
                                    return True
                        
                        except:
                            pass
                    
                    logger.warning("'Accesos' NO encontrado en descendientes")
                
                except Exception as e:
                    logger.error(f"Error iterando descendientes: {e}")
            
            logger.warning("No se pudo encontrar 'Accesos'")
            return False
        
        except Exception as e:
            logger.error(f"Error en _click_accesos_direct: {str(e)}")
            return False
    
    def expediente_loop_iei(self):
        """
        Bucle de búsqueda de expedientes FLEXIBLE que detecta el tipo de medida en cada expediente.
        
        IMPORTANTE: Después de procesar el primer expediente (IEI o DSE),
        el campo "Cambio de Expediente" está listo para ingresar el siguiente expediente.
        Este bucle continúa desde ese punto.
        
        CARACTERÍSTICA CLAVE: Detecta dinámicamente el TIPO DE MEDIDA de CADA expediente
        y ejecuta el flujo correspondiente (IEI o DSE).
        
        Flujo para cada expediente:
        1. Lee el siguiente expediente del Excel
        2. Ingresa el expediente
        3. Presiona ENTER para validar
        4. Verifica si es válido (sin error)
        5. Si es válido: 
           - Presiona ALT+A para validar ejecutor
           - Lee el TIPO DE MEDIDA de ese expediente
           - Si es IEI: Hace clic en "Trabar Intervención en Información"fill_interventor_and_plazo_loop()
           - Si es DSE: Hace clic en "Trabar Depósito sin Extracción"fill_monto_loop()
        6. Si es inválido: Marca como "EXP. INVALIDO" y continúa
        7. Repite hasta procesar todos los expedientes
        
        NOTA: El tipo de medida puede cambiar entre expedientes (ej: IEIDSEIEI)
        NOTA: Estos deben ser de la MISMA DEPENDENCIA (21 o 23)
        """
        logger.info("\n" + "=" * 70)
        logger.info("INICIANDO BUCLE DE EXPEDIENTES FLEXIBLE (DETECTANDO TIPO POR EXPEDIENTE)")
        logger.info("=" * 70)
        
        try:
            # Cargar datos del Excel
            excel_file = SCRIPT_DIR / "EXPEDIENTES.xlsx"
            expedientes = pd.read_excel(excel_file, engine="openpyxl", dtype=str)
            
            logger.info(f"Total de expedientes en Excel: {len(expedientes)}")
            logger.info(f"Procesados hasta ahora: {self.primer_expediente_idx + 1} expediente(s)")
            logger.info(f"Expedientes restantes a procesar: {len(expedientes) - (self.primer_expediente_idx + 1)}")
            
            # Procesar expedientes restantes (comenzando desde el siguiente al primer expediente válido)
            for idx in range(self.primer_expediente_idx + 1, len(expedientes)):
                exp_actual = str(expedientes.iloc[idx]["EXPEDIENTE"]).strip()
                
                logger.info(f"\n{'=' * 70}")
                logger.info(f"EXPEDIENTE {idx + 1}/{len(expedientes)}: {exp_actual}")
                logger.info(f"{'=' * 70}")
                
                try:
                    # En este punto, el campo de "Cambio de Expediente" está listo
                    # para que ingresemos el siguiente expediente
                    
                    # ============================================================
                    # PASO 1: Digitar expediente
                    # ============================================================
                    logger.info("PASO 1: Digitando expediente...")
                    time.sleep(0.3)
                    pyautogui.write(exp_actual, interval=0.05)
                    time.sleep(0.3)
                    
                    # ============================================================
                    # PASO 2: Presionar ENTER
                    # ============================================================
                    logger.info("PASO 2: Presionando ENTER...")
                    pyautogui.press('return')
                    time.sleep(0.5)
                    
                    # ============================================================
                    # PASO 3: Verificar si el expediente es válido
                    # ============================================================
                    logger.info("PASO 3: Verificando si el expediente es válido...")
                    time.sleep(0.5)
                    
                    is_valid = self.check_expediente_error_screen()
                    
                    if not is_valid:
                        # Expediente inválido
                        logger.warning(f"✗ Expediente inválido: {exp_actual}")
                        
                        # Cerrar ventana de error
                        self.close_expediente_window()
                        time.sleep(0.5)
                        
                        # Marcar como inválido en Excel
                        self.mark_invalid_expediente_in_results(idx)
                        
                        # Limpiar campo y continuar con el siguiente
                        logger.info("Limpiando campo de expediente...")
                        pyautogui.hotkey('ctrl', 'backspace')
                        time.sleep(0.2)
                        
                        logger.info("Continuando con el siguiente expediente...")
                        continue
                    
                    # Expediente válido
                    logger.info(f" Expediente válido: {exp_actual}")
                    
                    # ============================================================
                    # PASO 4: Presionar ALT+A para validar ejecutor
                    # ============================================================
                    logger.info("PASO 4: Presionando ALT+A para validar ejecutor...")
                    pyautogui.hotkey('alt', 'a')
                    time.sleep(1)
                    
                    # ============================================================
                    # PASO 5: Validar que el expediente tenga todos los datos necesarios
                    # ============================================================
                    logger.info("PASO 5: Validando datos del expediente...")
                    es_valido, mensaje_error = self.validate_expediente_row(expedientes, idx)
                    
                    if not es_valido:
                        logger.warning(f"✗ Expediente incompleto: {mensaje_error}")
                        
                        # Marcar con el motivo específico en Excel
                        self.mark_invalid_expediente_in_results(idx, mensaje_error)
                        
                        # Limpiar campo y continuar
                        logger.info("Limpiando campo de expediente...")
                        pyautogui.hotkey('ctrl', 'backspace')
                        time.sleep(0.2)
                        
                        logger.info("Continuando con el siguiente expediente...")
                        continue
                    
                    logger.info(f" Todos los datos del expediente son válidos")
                    
                    # ============================================================
                    # PASO 6: Leer el TIPO DE MEDIDA de la fila actual
                    # ============================================================
                    logger.info("PASO 6: Leyendo TIPO DE MEDIDA del expediente...")
                    tipo_medida_actual = str(expedientes.iloc[idx]["TIPO DE MEDIDA"]).strip().upper()
                    
                    # Detectar tipo (case-insensitive)
                    if "IEI" in tipo_medida_actual:
                        medida_tipo = "IEI"
                        logger.info(f" Tipo detectado: IEI")
                    elif "DSE" in tipo_medida_actual:
                        medida_tipo = "DSE"
                        logger.info(f" Tipo detectado: DSE")
                    else:
                        logger.error(f"Tipo no reconocido: {tipo_medida_actual}")
                        # Marcar como inválido y continuar
                        self.mark_invalid_expediente_in_results(idx)
                        continue
                    
                    # ============================================================
                    # PASO 6: Leer el TIPO DE MEDIDA de la fila actual
                    # ============================================================
                    logger.info("PASO 6: Leyendo TIPO DE MEDIDA del expediente...")
                    tipo_medida_actual = str(expedientes.iloc[idx]["TIPO DE MEDIDA"]).strip().upper()
                    
                    # Detectar tipo (case-insensitive)
                    if "IEI" in tipo_medida_actual:
                        medida_tipo = "IEI"
                        logger.info(f" Tipo detectado: IEI")
                    elif "DSE" in tipo_medida_actual:
                        medida_tipo = "DSE"
                        logger.info(f" Tipo detectado: DSE")
                    else:
                        logger.error(f"Tipo no reconocido: {tipo_medida_actual}")
                        # Marcar como inválido y continuar
                        self.mark_invalid_expediente_in_results(idx, f"TIPO NO VÁLIDO: {tipo_medida_actual}")
                        continue
                    
                    # ============================================================
                    # PASO 7: Hacer clic en "Trabar Embargo"
                    # ============================================================
                    logger.info("PASO 7: Haciendo clic en 'Trabar Embargo'...")
                    if not self.click_trabar_embargo():
                        logger.warning("No se pudo hacer clic en 'Trabar Embargo', continuando...")
                    time.sleep(0.5)
                    
                    # ============================================================
                    # PASO 8: Ejecutar flujo según el tipo detectado
                    # ============================================================
                    logger.info("PASO 8: Ejecutando flujo según tipo de medida...")
                    
                    if medida_tipo == "IEI":
                        logger.info("→ Ejecutando flujo IEI: 'Trabar Intervención en Información'")
                        
                        # Hacer clic en Trabar Intervención en Información
                        if not self.click_trabar_intervencion_informacion():
                            logger.error("No se pudo hacer clic en 'Trabar Intervención en Información'")
                            continue
                        
                        # Manejar aviso
                        logger.info("Manejando posible aviso...")
                        if not self.handle_trabar_intervencion_aviso():
                            logger.warning("No se pudo manejar el aviso, continuando...")
                        
                        # Rellenar INTERVENTOR y PLAZO
                        logger.info("Rellenando campos INTERVENTOR y PLAZO...")
                        if not self.fill_interventor_and_plazo_loop(idx):
                            logger.error(f"Error rellenando campos para expediente {idx + 1}")
                            continue
                    
                    elif medida_tipo == "DSE":
                        logger.info("→ Ejecutando flujo DSE: 'Trabar Depósito sin Extracción'")
                        
                        # Hacer clic en Trabar Depósito sin Extracción
                        if not self.click_trabar_deposito_sin_extraccion():
                            logger.error("No se pudo hacer clic en 'Trabar Depósito sin Extracción'")
                            continue
                        
                        # Manejar aviso
                        logger.info("Manejando posible aviso...")
                        if not self.handle_trabar_deposito_aviso():
                            logger.warning("No se pudo manejar el aviso, continuando...")
                        
                        # Rellenar MONTO
                        logger.info("Rellenando campo MONTO...")
                        if not self.fill_monto_loop(idx):
                            logger.error(f"Error rellenando MONTO para expediente {idx + 1}")
                            continue
                    
                    logger.info(f" Expediente {idx + 1} procesado correctamente")
                    logger.info(f"Campo de 'Cambio de Expediente' listo para el siguiente expediente")
                
                except Exception as e:
                    logger.error(f"Error procesando expediente {idx + 1}: {e}")
                    # Continuar con el siguiente expediente
                    continue
            
            logger.info("\n" + "=" * 70)
            logger.info(" BUCLE DE EXPEDIENTES FLEXIBLE COMPLETADO")
            logger.info("=" * 70)
            return True
        
        except Exception as e:
            logger.error(f"Error en expediente_loop_iei: {str(e)}")
            return False
    
    def expediente_loop(self):
        """
        Bucle de búsqueda de expedientes para IEI (Dependencia 21 - PRICO).
        
        Este bucle se ejecuta después de navegar a "Accesos".
        Flujo:
        1. Hacer doble clic en "Cambio de Expediente"
        2. Lee el siguiente expediente del Excel
        3. Ingresa el expediente
        4. Presiona ALT+A para validar
        5. Realiza el proceso IEI (Trabar Intervención en Información)
        6. Regresa al menú y repite
        
        El bucle continúa hasta que se procesen todos los expedientes en el Excel.
        """
        logger.info("\n" + "=" * 70)
        logger.info("INICIANDO BUCLE DE BÚSQUEDA DE EXPEDIENTES (IEI)")
        logger.info("=" * 70)
        
        try:
            # Cargar datos del Excel
            excel_file = SCRIPT_DIR / "EXPEDIENTES.xlsx"
            expedientes = pd.read_excel(excel_file, engine="openpyxl", dtype=str)
            
            logger.info(f"Total de expedientes a procesar: {len(expedientes)}")
            
            # Procesar cada expediente en el Excel (comenzando desde el siguiente al primer expediente válido)
            logger.info(f"Primer expediente válido encontrado en posición {self.primer_expediente_idx} (fila {self.primer_expediente_idx + 2})")
            logger.info(f"El bucle comenzará desde posición {self.primer_expediente_idx + 1} (fila {self.primer_expediente_idx + 3})")
            
            for idx in range(self.primer_expediente_idx + 1, len(expedientes)):
                logger.info("\n" + "=" * 70)
                logger.info(f"PROCESANDO EXPEDIENTE {idx + 1} de {len(expedientes)}")
                logger.info("=" * 70)
                
                # Obtener expediente y datos
                exp_actual = str(expedientes.iloc[idx]["EXPEDIENTE"]).strip()
                logger.info(f"Expediente: {exp_actual}")
                
                # Esperar a que cargue el menú
                time.sleep(1)
                
                # PASO 1: Hacer doble clic en "Cambio de Expediente"
                logger.info("=" * 70)
                logger.info("PASO 1: Haciendo doble clic en 'Cambio de Expediente'")
                logger.info("=" * 70)
                if not self.click_cambio_expediente():
                    logger.warning("No se pudo hacer clic en 'Cambio de Expediente', continuando...")
                
                # PASO 2: Ingresar el expediente
                logger.info("=" * 70)
                logger.info("PASO 2: Ingresando expediente")
                logger.info("=" * 70)
                logger.info(f"Digitando expediente: '{exp_actual}'")
                time.sleep(0.5)
                pyautogui.write(exp_actual, interval=0.05)
                logger.info(" Expediente ingresado")
                time.sleep(0.5)
                
                # PASO 3: Presionar ENTER para validar expediente
                logger.info("=" * 70)
                logger.info("PASO 3: Presionando ENTER para validar expediente")
                logger.info("=" * 70)
                pyautogui.press('return')
                time.sleep(1)
                
                # PASO 4: Detectar si expediente es válido o inválido
                logger.info("=" * 70)
                logger.info("PASO 4: Verificando si expediente es válido")
                logger.info("=" * 70)
                error_detected, error_message = self.detect_expediente_error(timeout=2)
                
                if error_detected:
                    logger.warning(f"✗ Expediente inválido: {error_message}")
                    logger.info("Guardando 'EXP. INVALIDO' en Excel...")
                    self.update_excel_result_for_row(idx, "EXP. INVALIDO")
                    
                    # Presionar ENTER para cerrar el diálogo de error
                    logger.info("Presionando ENTER para cerrar el diálogo de error...")
                    pyautogui.press('return')
                    time.sleep(0.5)
                    
                    # Borrar el expediente que quedó en el campo
                    logger.info("Borrando el expediente del campo...")
                    pyautogui.hotkey('ctrl', 'a')  # Seleccionar todo
                    time.sleep(0.2)
                    pyautogui.press('delete')  # Borrar
                    time.sleep(0.5)
                    
                    logger.info("Pasando al siguiente expediente...")
                    continue
                
                logger.info(" Expediente válido")
                
                # PASO 5: Presionar ALT+A para validar y pasar al formulario
                logger.info("=" * 70)
                logger.info("PASO 5: Presionando ALT+A para validar")
                logger.info("=" * 70)
                pyautogui.hotkey('alt', 'a')
                time.sleep(1)
                logger.info(" ALT+A presionado")
                
                # PASO 6: Rellenar campos INTERVENTOR y PLAZO (sin navegación extra, directo en el formulario)
                logger.info("=" * 70)
                logger.info("PASO 6: Rellenando campos INTERVENTOR y PLAZO...")
                logger.info("=" * 70)
                
                # Cargar datos del Excel
                excel_file = SCRIPT_DIR / "EXPEDIENTES.xlsx"
                expedientes_data = pd.read_excel(excel_file, engine="openpyxl", dtype=str)
                
                # Obtener valores de INTERVENTOR y PLAZO de la fila actual
                interventor = None
                plazo = None
                
                if "INTERVENTOR" in expedientes_data.columns:
                    interventor = str(expedientes_data.iloc[idx]["INTERVENTOR"]).strip()
                else:
                    logger.error("Columna INTERVENTOR no encontrada")
                    continue
                
                if "PLAZO" in expedientes_data.columns:
                    plazo = str(expedientes_data.iloc[idx]["PLAZO"]).strip()
                else:
                    logger.error("Columna PLAZO no encontrada")
                    continue
                
                # Esperar 0.5 segundos y digitar INTERVENTOR
                logger.info(f"Digitando INTERVENTOR: '{interventor}'")
                time.sleep(0.5)
                pyautogui.write(interventor, interval=0.05)
                time.sleep(0.2)
                
                # Presionar TAB para pasar a PLAZO
                logger.info("Presionando TAB para pasar a PLAZO...")
                pyautogui.press('tab')
                time.sleep(0.2)
                
                # Digitar PLAZO
                logger.info(f"Digitando PLAZO: '{plazo}'")
                pyautogui.write(plazo, interval=0.05)
                time.sleep(0.3)
                
                # Presionar ALT+A para confirmar INTERVENTOR y PLAZO
                logger.info("Presionando ALT+A para confirmar INTERVENTOR y PLAZO...")
                pyautogui.hotkey('alt', 'a')
                time.sleep(1)
                
                # Detectar mensaje de aviso
                logger.info("Detectando posible mensaje de aviso...")
                aviso_detected, aviso_mensaje = self.detect_expediente_aviso(timeout=2)
                
                if aviso_detected:
                    logger.info(f"Aviso detectado: {aviso_mensaje[:100]}")
                
                # Presionar ALT+S dos veces para cerrar avisos y continuar
                logger.info("Presionando ALT+S...")
                time.sleep(0.5)
                pyautogui.hotkey('alt', 's')
                time.sleep(1)
                
                logger.info("Presionando ALT+S (2do)...")
                pyautogui.hotkey('alt', 's')
                time.sleep(1)
                
                # Detectar y extraer RC
                logger.info("Detectando mensaje de Resolución Coactiva...")
                rc_detected, rc_mensaje = self.detect_resolucion_coactiva_aviso(timeout=2)
                
                if rc_detected:
                    rc_number = self.extract_resolucion_coactiva_number(rc_mensaje)
                    logger.info(f"RC extraído: {rc_number}")
                    
                    # Guardar en Excel
                    self.update_excel_result_for_row(idx, rc_number)
                else:
                    logger.warning("No se detectó RC")
                
                # Presionar ENTER para aceptar
                logger.info("Presionando ENTER...")
                pyautogui.press('return')
                time.sleep(1)
                
                # Presionar ALT+C para regresar al menú
                logger.info("Presionando ALT+C para regresar al menú...")
                time.sleep(0.5)
                pyautogui.hotkey('alt', 'c')
                time.sleep(1)
                logger.info(" ALT+C presionado")
                
                logger.info(f" Expediente {idx + 1} procesado exitosamente")
            
            logger.info("\n" + "=" * 70)
            logger.info(" BUCLE DE EXPEDIENTES COMPLETADO")
            logger.info("=" * 70)
            return True
        
        except Exception as e:
            logger.error(f"Error en expediente_loop: {str(e)}")
            return False
    
    def fill_interventor_and_plazo_loop(self, row_idx):
        """
        Versión del relleno de INTERVENTOR y PLAZO para usar dentro del bucle de expedientes.
        Lee los datos de la fila especificada en el Excel.
        
        Args:
            row_idx: Índice de fila del Excel a procesar
        """
        logger.info(f"Rellenando campos INTERVENTOR y PLAZO para expediente en fila {row_idx + 1}...")
        
        try:
            # Cargar datos del Excel
            excel_file = SCRIPT_DIR / "EXPEDIENTES.xlsx"
            expedientes = pd.read_excel(excel_file, engine="openpyxl", dtype=str)
            
            # Obtener valores de INTERVENTOR y PLAZO de la fila actual
            interventor = None
            plazo = None
            
            if "INTERVENTOR" in expedientes.columns:
                interventor = str(expedientes.iloc[row_idx]["INTERVENTOR"]).strip()
            else:
                logger.error("Columna INTERVENTOR no encontrada")
                return False
            
            if "PLAZO" in expedientes.columns:
                plazo = str(expedientes.iloc[row_idx]["PLAZO"]).strip()
            else:
                logger.error("Columna PLAZO no encontrada")
                return False
            
            # Esperar 0.5 segundos y digitar INTERVENTOR
            logger.info(f"Digitando INTERVENTOR: '{interventor}'")
            time.sleep(0.5)
            pyautogui.write(interventor, interval=0.05)
            time.sleep(0.2)
            
            # Presionar TAB para pasar a PLAZO
            logger.info("Presionando TAB para pasar a PLAZO...")
            pyautogui.press('tab')
            time.sleep(0.2)
            
            # Digitar PLAZO
            logger.info(f"Digitando PLAZO: '{plazo}'")
            pyautogui.write(plazo, interval=0.05)
            time.sleep(0.3)
            
            # Presionar ALT+A para confirmar
            logger.info("Presionando ALT+A para confirmar...")
            pyautogui.hotkey('alt', 'a')
            time.sleep(1)
            
            # Detectar mensaje de aviso
            logger.info("Detectando posible mensaje de aviso...")
            aviso_detected, aviso_mensaje = self.detect_expediente_aviso(timeout=2)
            
            if aviso_detected:
                logger.info(f"Aviso detectado: {aviso_mensaje[:100]}")
            
            # Presionar ALT+S dos veces
            logger.info("Presionando ALT+S...")
            time.sleep(0.5)
            pyautogui.hotkey('alt', 's')
            time.sleep(1)
            
            logger.info("Presionando ALT+S (2do)...")
            pyautogui.hotkey('alt', 's')
            time.sleep(1)
            
            # Detectar y extraer RC
            logger.info("Detectando mensaje de Resolución Coactiva...")
            rc_detected, rc_mensaje = self.detect_resolucion_coactiva_aviso(timeout=2)
            
            if rc_detected:
                rc_number = self.extract_resolucion_coactiva_number(rc_mensaje)
                logger.info(f"RC extraído: {rc_number}")
                
                # Guardar en Excel
                self.update_excel_result_for_row(row_idx, rc_number)
            
            # Presionar ENTER y ALT+C para regresar
            logger.info("Presionando ENTER...")
            pyautogui.press('return')
            time.sleep(1)
            
            logger.info("Presionando ALT+C para regresar al menú...")
            time.sleep(0.5)
            pyautogui.hotkey('alt', 'c')
            time.sleep(1)
            
            # ================================================================
            # FLUJO POST-RC: Eliminar desplazamientos y preparar para siguiente expediente
            # ================================================================
            logger.info("=" * 70)
            logger.info("FLUJO POST-RC: Eliminando desplazamientos del menú")
            logger.info("=" * 70)
            
            # PASO 1: Hacer clic en "Trabar Embargo" para eliminar desplazamiento 1
            logger.info("Haciendo clic en 'Trabar Embargo' (eliminar desplazamiento 1)...")
            if not self.click_trabar_embargo():
                logger.warning("No se pudo hacer clic en 'Trabar Embargo'")
            time.sleep(0.5)
            
            # PASO 2: Hacer clic en "Proceso de Embargo" para eliminar desplazamiento 2
            logger.info("Haciendo clic en 'Proceso de Embargo' (eliminar desplazamiento 2)...")
            if not self.click_proceso_embargo():
                logger.warning("No se pudo hacer clic en 'Proceso de Embargo'")
            time.sleep(0.5)
            
            # PASO 3: Hacer clic en "Accesos" (ahora visible)
            logger.info("Haciendo clic en 'Accesos' (sin desplazamiento previo)...")
            if not self._click_accesos_direct():
                logger.warning("No se pudo hacer clic en 'Accesos'")
            time.sleep(0.5)
            
            # PASO 4: Hacer doble clic en "Cambio de Expediente"
            logger.info("Haciendo doble clic en 'Cambio de Expediente'...")
            if not self.click_cambio_expediente():
                logger.warning("No se pudo hacer clic en 'Cambio de Expediente'")
            time.sleep(0.5)
            
            logger.info("=" * 70)
            logger.info(" Campos completados exitosamente")
            logger.info("= Listo para siguiente expediente =")
            logger.info("=" * 70)
            return True
        
        except Exception as e:
            logger.error(f"Error en fill_interventor_and_plazo_loop: {str(e)}")
            return False
    
    def detect_desea_continuar_aviso(self, timeout=2):
        """
        Detecta si aparece el mensaje "¿ Desea Continuar ?" (con espacios).
        Este aviso es específico de DSE cuando hay embargos activos.
        
        Usa MSAA para inspeccionar controles de tipo Text.
        
        Retorna:
            - Tupla (True, mensaje) si se detecta el mensaje
            - Tupla (False, "") si no hay mensaje
        """
        try:
            logger.info("Verificando si hay mensaje '¿ Desea Continuar ?'...")
            desktop = Desktop(backend="uia")
            end_time = time.time() + timeout
            
            while time.time() < end_time:
                try:
                    # Buscar ventanas de diálogo
                    dlgs = desktop.windows()
                    
                    for dlg in dlgs:
                        try:
                            # Buscar controles de tipo Text en descendientes
                            descendants = dlg.descendants()
                            
                            for desc in descendants:
                                try:
                                    texto = desc.window_text()
                                    
                                    # Búsqueda ROBUSTA: múltiples palabras clave
                                    # Patrón: "¿ Desea Continuar ?" - requiere AMBAS palabras
                                    texto_lower = texto.lower()
                                    if "desea" in texto_lower and "continuar" in texto_lower:
                                        logger.warning(f"Aviso '¿ Desea Continuar ?' detectado")
                                        return (True, texto)
                                
                                except Exception:
                                    pass
                        
                        except Exception:
                            pass
                
                except Exception:
                    pass
                
                time.sleep(0.1)
            
            return (False, "")
        
        except Exception as e:
            logger.error(f"Error en detect_desea_continuar_aviso: {e}")
            return (False, "")
    
    def detect_grabar_resolucion_aviso(self, timeout=2):
        """
        Detecta si aparece el mensaje "¿Desea Ud. grabar la Resolución Coactiva?"
        Este aviso es opcional y puede aparecer en DSE.
        
        Usa MSAA para inspeccionar controles de tipo Text.
        
        Retorna:
            - Tupla (True, mensaje) si se detecta el mensaje
            - Tupla (False, "") si no hay mensaje
        """
        try:
            logger.info("Verificando si hay mensaje '¿Desea Ud. grabar la Resolución?'...")
            desktop = Desktop(backend="uia")
            end_time = time.time() + timeout
            
            while time.time() < end_time:
                try:
                    # Buscar ventanas de diálogo
                    dlgs = desktop.windows()
                    
                    for dlg in dlgs:
                        try:
                            # Buscar controles de tipo Text en descendientes
                            descendants = dlg.descendants()
                            
                            for desc in descendants:
                                try:
                                    texto = desc.window_text()
                                    
                                    # Búsqueda ROBUSTA: múltiples palabras clave
                                    # Patrón: "¿Desea Ud. grabar la Resolución Coactiva?"
                                    # Requiere: "desea" + "grabar" + ("resolucion" OR "coactiva")
                                    texto_lower = texto.lower()
                                    if ("desea" in texto_lower and 
                                        "grabar" in texto_lower and 
                                        ("resolucion" in texto_lower or "coactiva" in texto_lower)):
                                        logger.warning(f"Aviso '¿Desea Ud. grabar Resolución?' detectado")
                                        return (True, texto)
                                
                                except Exception:
                                    pass
                        
                        except Exception:
                            pass
                
                except Exception:
                    pass
                
                time.sleep(0.1)
            
            return (False, "")
        
        except Exception as e:
            logger.error(f"Error en detect_grabar_resolucion_aviso: {e}")
            return (False, "")
    
    def fill_monto(self):
        """
        Llena el campo de MONTO en el formulario (DSE - MEPECO).
        
        Flujo para MONTO MAYOR:
        1. Digita MONTO
        2. Presiona ALT+A
        3. Detecta: "El monto ingresado excede en más del X% el Saldo del Expediente"
        4. Guarda "MONTO MAYOR" en Excel
        5. Presiona ENTER
        6. Detecta: "El monto de embargo ingresado supera el saldo embargable del expediente..."
        7. Presiona ENTER para continuar
       FIN (Se queda a la espera del bucle de expedientes)
        
        Flujo para MONTO ACEPTADO:
        1. Digita MONTO
        2. Presiona ALT+A
        3. Detecta (OPCIONAL): "El Expediente XXX correspondiente al RUC XXX tiene X Embargos activos"
          Si aparece: ALT+S
        4. Detecta (OPCIONAL): "¿ Desea Continuar ?" (con espacios específicos)
          Si aparece: ALT+S
        5. Detecta (OPCIONAL): "¿Desea Ud. grabar la Resolución Coactiva?"
          Si aparece: ALT+S
        6. Detecta: "Se grabó la Resolución Coactiva con el número: XXX"
          Extrae RC del mensaje
          Guarda RC en Excel (en columna RESULTADO)
          Presiona ALT+A para cerrar
       FIN (Se queda a la espera del bucle de expedientes)
        """
        logger.info("Rellenando campo de MONTO (DSE)...")
        
        try:
            # Cargar datos del Excel
            excel_file = SCRIPT_DIR / "EXPEDIENTES.xlsx"
            expedientes = pd.read_excel(excel_file, engine="openpyxl", dtype=str)
            
            # Obtener valor de MONTO del expediente actual (self.primer_expediente_idx)
            monto = None
            
            if "MONTO" in expedientes.columns:
                monto = str(expedientes.iloc[self.primer_expediente_idx]["MONTO"]).strip()
                logger.info(f"MONTO obtenido del Excel (fila {self.primer_expediente_idx + 2}): '{monto}'")
            else:
                logger.warning("No existe columna 'MONTO' en el Excel")
                monto = ""
            
            # ============================================================
            # PASO 1: Esperar 0.5s y digitar MONTO
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 1: Esperando 0.5s y digitando MONTO")
            logger.info("=" * 70)
            
            logger.info("Esperando 0.5 segundos...")
            time.sleep(0.5)
            
            logger.info(f"Digitando MONTO: '{monto}'")
            pyautogui.write(monto, interval=0.05)
            time.sleep(0.3)
            
            # ============================================================
            # PASO 2: Presionar ALT+A para confirmar MONTO
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 2: Presionando ALT+A para confirmar MONTO")
            logger.info("=" * 70)
            
            logger.info("Presionando ALT+A...")
            pyautogui.hotkey('alt', 'a')
            time.sleep(1)
            logger.info(" ALT+A presionado correctamente")
            
            # ============================================================
            # PASO 3: Detectar si es MONTO MAYOR
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 3: Detectando si es MONTO MAYOR")
            logger.info("=" * 70)
            
            monto_mayor_detectado, monto_mayor_msg = self.detect_monto_aviso(timeout=2)
            
            if monto_mayor_detectado:
                logger.warning(f"MONTO MAYOR detectado: {monto_mayor_msg}")
                
                # Guardar "MONTO MAYOR" en Excel
                logger.info("Guardando 'MONTO MAYOR' en Excel...")
                self.update_excel_result("MONTO MAYOR")
                logger.info(" 'MONTO MAYOR' guardado en Excel")
                
                # Presionar ENTER para cerrar el aviso de MONTO MAYOR
                logger.info("Presionando ENTER para cerrar aviso de MONTO MAYOR...")
                pyautogui.press('return')
                time.sleep(1)
                logger.info(" ENTER presionado")
                
                # Detectar el segundo aviso de MONTO MAYOR
                logger.info("Detectando segundo aviso: 'El monto de embargo ingresado supera el saldo...'")
                time.sleep(0.5)
                segundo_aviso_detected, segundo_aviso_msg = self.detect_monto_aviso(timeout=2)
                
                if segundo_aviso_detected:
                    logger.warning(f"Segundo aviso detectado: {segundo_aviso_msg}")
                    logger.info("Presionando ENTER para cerrar segundo aviso...")
                    pyautogui.press('return')
                    time.sleep(1)
                    logger.info(" ENTER presionado")
                
                logger.info("=" * 70)
                logger.info(" FLUJO DE MONTO MAYOR - INICIANDO BUCLE PARA SIGUIENTE EXPEDIENTE")
                logger.info("=" * 70)
                
                # ============================================================
                # PASO 4: Presionar ALT+C para regresar al menú
                # ============================================================
                logger.info("=" * 70)
                logger.info("PASO 4: Presionando ALT+C para regresar al menú")
                logger.info("=" * 70)
                
                logger.info("Presionando ALT+C para regresar al menú...")
                time.sleep(0.5)
                pyautogui.hotkey('alt', 'c')
                time.sleep(1)
                logger.info(" ALT+C presionado correctamente")
                
                # ============================================================
                # PASO 5: Eliminar desplazamientos del menú para acceder a "Accesos"
                # ============================================================
                logger.info("=" * 70)
                logger.info("PASO 5: Eliminando desplazamientos del menú")
                logger.info("=" * 70)
                
                # Clic en "Trabar Embargo" para eliminar un desplazamiento
                logger.info("Haciendo clic en 'Trabar Embargo' para eliminar desplazamiento...")
                if not self.click_trabar_embargo():
                    logger.warning("No se pudo hacer clic en 'Trabar Embargo'")
                time.sleep(0.5)
                
                # Clic en "Proceso de Embargo" para eliminar otro desplazamiento
                logger.info("Haciendo clic en 'Proceso de Embargo' para eliminar desplazamiento...")
                if not self.click_proceso_embargo():
                    logger.warning("No se pudo hacer clic en 'Proceso de Embargo'")
                time.sleep(0.5)
                
                logger.info(" Desplazamientos eliminados - 'Accesos' ahora visible en el mismo nivel")
                
                # ============================================================
                # PASO 6: Hacer clic en "Accesos"
                # ============================================================
                logger.info("=" * 70)
                logger.info("PASO 6: Haciendo clic en 'Accesos'")
                logger.info("=" * 70)
                
                # Buscar y hacer clic en "Accesos" directamente (sin desplazamiento)
                logger.info("Buscando 'Accesos' en el menú...")
                if not self._click_accesos_direct():
                    logger.warning("No se pudo hacer clic en 'Accesos'")
                
                time.sleep(0.5)
                
                # ============================================================
                # PASO 7: Hacer doble clic en "Cambio de Expediente"
                # ============================================================
                logger.info("=" * 70)
                logger.info("PASO 7: Haciendo doble clic en 'Cambio de Expediente'")
                logger.info("=" * 70)
                
                if not self.click_cambio_expediente():
                    logger.warning("No se pudo hacer clic en 'Cambio de Expediente'")
                
                time.sleep(0.5)
                
                # ============================================================
                # PASO 8: Iniciar bucle de expedientes
                # ============================================================
                logger.info("=" * 70)
                logger.info("PASO 8: Iniciando bucle de búsqueda de expedientes restantes")
                logger.info("=" * 70)
                
                if not self.expediente_loop_dse():
                    logger.warning("Error en expediente_loop_dse")
                
                logger.info(" PROCESO COMPLETADO EXITOSAMENTE")
                return True
            
            # ============================================================
            # PASO 4: MONTO ACEPTADO - Detectar avisos opcionales
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 4: MONTO ACEPTADO - Detectando avisos opcionales")
            logger.info("=" * 70)
            
            # Aviso OPCIONAL 1: "El Expediente XXX correspondiente al RUC XXX tiene X Embargos activos..."
            logger.info("Detectando aviso de embargos activos (OPCIONAL)...")
            embargo_detectado, embargo_msg = self.detect_expediente_aviso(timeout=2)
            
            if embargo_detectado:
                logger.warning(f"Aviso de embargos detectado: {embargo_msg}")
                logger.info("Presionando ALT+S...")
                pyautogui.hotkey('alt', 's')
                time.sleep(1)
                logger.info(" ALT+S presionado")
            else:
                logger.info("No se detectó aviso de embargos (continuando...)")
            
            # Aviso OPCIONAL 2: "¿ Desea Continuar ?" (con espacios específicos)
            logger.info("Detectando aviso '¿ Desea Continuar ?' (OPCIONAL)...")
            desea_continuar_detectado, desea_continuar_msg = self.detect_desea_continuar_aviso(timeout=2)
            
            if desea_continuar_detectado:
                logger.warning(f"Aviso '¿ Desea Continuar ?' detectado: {desea_continuar_msg}")
                logger.info("Presionando ALT+S...")
                pyautogui.hotkey('alt', 's')
                time.sleep(1)
                logger.info(" ALT+S presionado")
            else:
                logger.info("No se detectó '¿ Desea Continuar ?' (continuando...)")
            
            # ============================================================
            # PASO 5: Detectar "¿Desea Ud. grabar la Resolución?" (OPCIONAL)
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 5: Detectando '¿Desea Ud. grabar la Resolución Coactiva?' (OPCIONAL)")
            logger.info("=" * 70)
            
            grabar_resolucion_detectado, grabar_resolucion_msg = self.detect_grabar_resolucion_aviso(timeout=2)
            
            if grabar_resolucion_detectado:
                logger.warning(f"Aviso '¿Desea Ud. grabar Resolución?' detectado: {grabar_resolucion_msg}")
                logger.info("Presionando ALT+S...")
                pyautogui.hotkey('alt', 's')
                time.sleep(1)
                logger.info(" ALT+S presionado")
            else:
                logger.info("No se detectó '¿Desea Ud. grabar Resolución?' (continuando...)")
            
            # ============================================================
            # PASO 6: Detectar y extraer Resolución Coactiva (OBLIGATORIO)
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 6: Detectando mensaje de Resolución Coactiva (OBLIGATORIO)")
            logger.info("=" * 70)
            
            rc_detectado, rc_msg = self.detect_resolucion_coactiva_aviso(timeout=2)
            
            if rc_detectado:
                logger.warning(f"RC detectado: {rc_msg}")
                
                # Extraer número de RC
                rc_number = self.extract_resolucion_coactiva_number(rc_msg)
                
                if rc_number:
                    logger.info(f"RC extraído: {rc_number}")
                    logger.info("Guardando RC en Excel...")
                    self.update_excel_result(rc_number)
                    logger.info(" RC guardado en Excel")
                else:
                    logger.warning("No se pudo extraer RC del mensaje")
                    self.update_excel_result("RC NO EXTRAÍDO")
                
                # Presionar ENTER para cerrar el mensaje de RC
                logger.info("Presionando ENTER para cerrar mensaje de RC...")
                pyautogui.press('return')
                time.sleep(1)
                logger.info(" ENTER presionado")
            else:
                logger.warning("No se detectó mensaje de RC")
                self.update_excel_result("RC NO DETECTADO")
            
            logger.info("=" * 70)
            logger.info(" FLUJO DE MONTO ACEPTADO - INICIANDO BUCLE PARA SIGUIENTE EXPEDIENTE")
            logger.info("=" * 70)
            
            # ============================================================
            # PASO 7: Presionar ALT+C para regresar al menú
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 7: Presionando ALT+C para regresar al menú")
            logger.info("=" * 70)
            
            logger.info("Presionando ALT+C para regresar al menú...")
            time.sleep(0.5)
            pyautogui.hotkey('alt', 'c')
            time.sleep(1)
            logger.info(" ALT+C presionado correctamente")
            
            # ============================================================
            # PASO 8: Eliminar desplazamientos del menú para acceder a "Accesos"
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 8: Eliminando desplazamientos del menú")
            logger.info("=" * 70)
            
            # Clic en "Trabar Embargo" para eliminar un desplazamiento
            logger.info("Haciendo clic en 'Trabar Embargo' para eliminar desplazamiento...")
            if not self.click_trabar_embargo():
                logger.warning("No se pudo hacer clic en 'Trabar Embargo'")
            time.sleep(0.5)
            
            # Clic en "Proceso de Embargo" para eliminar otro desplazamiento
            logger.info("Haciendo clic en 'Proceso de Embargo' para eliminar desplazamiento...")
            if not self.click_proceso_embargo():
                logger.warning("No se pudo hacer clic en 'Proceso de Embargo'")
            time.sleep(0.5)
            
            logger.info(" Desplazamientos eliminados - 'Accesos' ahora visible en el mismo nivel")
            
            # ============================================================
            # PASO 9: Hacer clic en "Accesos"
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 9: Haciendo clic en 'Accesos'")
            logger.info("=" * 70)
            
            # Buscar y hacer clic en "Accesos" directamente (sin desplazamiento)
            logger.info("Buscando 'Accesos' en el menú...")
            if not self._click_accesos_direct():
                logger.warning("No se pudo hacer clic en 'Accesos'")
            
            time.sleep(0.5)
            
            # ============================================================
            # PASO 10: Hacer doble clic en "Cambio de Expediente"
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 10: Haciendo doble clic en 'Cambio de Expediente'")
            logger.info("=" * 70)
            
            if not self.click_cambio_expediente():
                logger.warning("No se pudo hacer clic en 'Cambio de Expediente'")
            
            time.sleep(0.5)
            
            # ============================================================
            # PASO 11: Iniciar bucle de expedientes
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 11: Iniciando bucle de búsqueda de expedientes restantes")
            logger.info("=" * 70)
            
            if not self.expediente_loop_dse():
                logger.warning("Error en expediente_loop_dse")
            
            logger.info(" PROCESO COMPLETADO EXITOSAMENTE")
            return True
        
        except Exception as e:
            logger.error(f"Error en fill_monto: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return False
    
    def update_excel_executor_result(self, resultado):
        """
        Actualiza el Excel con el resultado de la validación del ejecutor.
        Agrega o actualiza la columna RESULTADO en la fila actual.
        """
        try:
            excel_file = SCRIPT_DIR / "EXPEDIENTES.xlsx"
            
            # Leer el Excel preservando formato original
            expedientes = pd.read_excel(excel_file, engine="openpyxl", dtype=str)
            
            # Crear columna si no existe
            if "RESULTADO" not in expedientes.columns:
                logger.info("Creando columna 'RESULTADO' en el Excel")
                expedientes["RESULTADO"] = None
            
            # Actualizar la primera fila (la que procesamos)
            expedientes.at[0, "RESULTADO"] = resultado
            
            # Guardar el Excel preservando formato
            expedientes.to_excel(excel_file, engine="openpyxl", index=False)
            logger.info(f"Excel actualizado: RESULTADO = '{resultado}'")
            # Si se escribió un RC o MONTO MAYOR, marcar como completado
            try:
                if isinstance(resultado, str):
                    res_up = resultado.strip().upper()
                    if resultado.isdigit() or res_up == 'MONTO MAYOR':
                        self.last_exp_completed = True
                    else:
                        self.last_exp_completed = False
                else:
                    self.last_exp_completed = False
            except Exception:
                pass
            
            return True
        
        except Exception as e:
            logger.error(f"Error actualizando Excel: {str(e)}")
            return False
    
    def update_excel_result_for_row(self, row_idx, resultado):
        """
        Actualiza el Excel con el resultado para una fila específica.
        Usado en el bucle de expedientes para guardar RC de cada expediente.
        
        Args:
            row_idx: Índice de fila (0-based)
            resultado: Valor a guardar en la columna RESULTADO
        """
        try:
            logger.info(f"Actualizando Excel para fila {row_idx + 1} con resultado: {resultado}")
            
            excel_file = SCRIPT_DIR / "EXPEDIENTES.xlsx"
            from openpyxl import load_workbook
            
            # Leer el archivo con openpyxl para preservar formatos
            wb = load_workbook(excel_file)
            ws = wb.active
            
            # Encontrar la columna RESULTADO
            headers = {}
            for col_idx, cell in enumerate(ws[1], 1):
                headers[cell.value] = col_idx
            
            # Si no existe la columna RESULTADO, crearla
            if "RESULTADO" not in headers:
                resultado_col = len(headers) + 1
                ws.cell(row=1, column=resultado_col, value="RESULTADO")
            else:
                resultado_col = headers["RESULTADO"]
            
            # Escribir el resultado en la fila especificada (sumando 2 porque: 1 es header, +1 para 0-based)
            celda_resultado = ws.cell(row=row_idx + 2, column=resultado_col, value=resultado)
            
            # Aplicar formato de texto si es un número con leading zero
            if isinstance(resultado, str) and resultado.isdigit():
                celda_resultado.number_format = '@'
                
                # Si comienza con 0, anteponer apóstrofe
                if resultado.startswith('0'):
                    celda_resultado.value = "'" + resultado
            
            # Guardar el archivo
            wb.save(excel_file)
            logger.info(f" Excel actualizado para fila {row_idx + 1}")
            # Si escribimos un RC (número) o un indicador de MONTO MAYOR,
            # consideramos que el expediente quedó completado y por lo
            # tanto habilitamos el cambio de expediente.
            try:
                if isinstance(resultado, str):
                    res_up = resultado.strip().upper()
                    if resultado.isdigit() or res_up == 'MONTO MAYOR':
                        self.last_exp_completed = True
                    else:
                        # Otros mensajes no implican completar el flujo
                        self.last_exp_completed = False
                else:
                    self.last_exp_completed = False
            except Exception:
                pass
            return True
        
        except Exception as e:
            logger.error(f"Error actualizando Excel para fila {row_idx + 1}: {e}")
            return False
    
    def fill_monto_loop(self, row_idx):
        """
        Versión del relleno de MONTO para usar dentro del bucle de expedientes DSE.
        Lee los datos de la fila especificada en el Excel.
        
        Flujo idéntico a fill_monto() pero para una fila específica del Excel.
        
        Args:
            row_idx: Índice de fila del Excel a procesar
        """
        logger.info(f"Rellenando campo MONTO para expediente en fila {row_idx + 1}...")
        
        try:
            # Cargar datos del Excel
            excel_file = SCRIPT_DIR / "EXPEDIENTES.xlsx"
            expedientes = pd.read_excel(excel_file, engine="openpyxl", dtype=str)
            
            # Obtener valor de MONTO de la fila especificada
            monto = None
            
            if "MONTO" in expedientes.columns:
                monto = str(expedientes.iloc[row_idx]["MONTO"]).strip()
                logger.info(f"MONTO obtenido del Excel (fila {row_idx + 2}): '{monto}'")
            else:
                logger.warning("No existe columna 'MONTO' en el Excel")
                monto = ""
            
            # ============================================================
            # PASO 1: Esperar 0.5s y digitar MONTO
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 1: Esperando 0.5s y digitando MONTO")
            logger.info("=" * 70)
            
            logger.info("Esperando 0.5 segundos...")
            time.sleep(0.5)
            
            logger.info(f"Digitando MONTO: '{monto}'")
            pyautogui.write(monto, interval=0.05)
            time.sleep(0.3)
            
            # ============================================================
            # PASO 2: Presionar ALT+A para confirmar MONTO
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 2: Presionando ALT+A para confirmar MONTO")
            logger.info("=" * 70)
            
            logger.info("Presionando ALT+A...")
            pyautogui.hotkey('alt', 'a')
            time.sleep(1)
            logger.info(" ALT+A presionado correctamente")
            
            # ============================================================
            # PASO 3: Detectar si es MONTO MAYOR
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 3: Detectando si es MONTO MAYOR")
            logger.info("=" * 70)
            
            monto_mayor_detectado, monto_mayor_msg = self.detect_monto_aviso(timeout=2)
            
            if monto_mayor_detectado:
                logger.warning(f"MONTO MAYOR detectado: {monto_mayor_msg}")
                
                # Guardar "MONTO MAYOR" en Excel
                logger.info("Guardando 'MONTO MAYOR' en Excel...")
                self.update_excel_result_for_row(row_idx, "MONTO MAYOR")
                logger.info(" 'MONTO MAYOR' guardado en Excel")
                
                # Presionar ENTER para cerrar el aviso de MONTO MAYOR
                logger.info("Presionando ENTER para cerrar aviso de MONTO MAYOR...")
                pyautogui.press('return')
                time.sleep(1)
                logger.info(" ENTER presionado")
                
                # Detectar el segundo aviso de MONTO MAYOR
                logger.info("Detectando segundo aviso: 'El monto de embargo ingresado supera el saldo...'")
                time.sleep(0.5)
                segundo_aviso_detected, segundo_aviso_msg = self.detect_monto_aviso(timeout=2)
                
                if segundo_aviso_detected:
                    logger.warning(f"Segundo aviso detectado: {segundo_aviso_msg}")
                    logger.info("Presionando ENTER para cerrar segundo aviso...")
                    pyautogui.press('return')
                    time.sleep(1)
                    logger.info(" ENTER presionado")
                
                logger.info("=" * 70)
                logger.info(" FLUJO DE MONTO MAYOR - INICIANDO BUCLE PARA SIGUIENTE EXPEDIENTE")
                logger.info("=" * 70)
                
                # ============================================================
                # PASO 4: Presionar ALT+C para regresar al menú
                # ============================================================
                logger.info("=" * 70)
                logger.info("PASO 4: Presionando ALT+C para regresar al menú")
                logger.info("=" * 70)
                
                logger.info("Presionando ALT+C para regresar al menú...")
                time.sleep(0.5)
                pyautogui.hotkey('alt', 'c')
                time.sleep(1)
                logger.info(" ALT+C presionado correctamente")
                
                # ============================================================
                # PASO 5: Eliminar desplazamientos del menú para acceder a "Accesos"
                # ============================================================
                logger.info("=" * 70)
                logger.info("PASO 5: Eliminando desplazamientos del menú")
                logger.info("=" * 70)
                
                # Clic en "Trabar Embargo" para eliminar un desplazamiento
                logger.info("Haciendo clic en 'Trabar Embargo' para eliminar desplazamiento...")
                if not self.click_trabar_embargo():
                    logger.warning("No se pudo hacer clic en 'Trabar Embargo'")
                time.sleep(0.5)
                
                # Clic en "Proceso de Embargo" para eliminar otro desplazamiento
                logger.info("Haciendo clic en 'Proceso de Embargo' para eliminar desplazamiento...")
                if not self.click_proceso_embargo():
                    logger.warning("No se pudo hacer clic en 'Proceso de Embargo'")
                time.sleep(0.5)
                
                logger.info(" Desplazamientos eliminados - 'Accesos' ahora visible en el mismo nivel")
                
                # ============================================================
                # PASO 6: Hacer clic en "Accesos"
                # ============================================================
                logger.info("=" * 70)
                logger.info("PASO 6: Haciendo clic en 'Accesos'")
                logger.info("=" * 70)
                
                # Buscar y hacer clic en "Accesos" directamente (sin desplazamiento)
                logger.info("Buscando 'Accesos' en el menú...")
                if not self._click_accesos_direct():
                    logger.warning("No se pudo hacer clic en 'Accesos'")
                
                time.sleep(0.5)
                
                # ============================================================
                # PASO 7: Hacer doble clic en "Cambio de Expediente"
                # ============================================================
                logger.info("=" * 70)
                logger.info("PASO 7: Haciendo doble clic en 'Cambio de Expediente'")
                logger.info("=" * 70)
                
                if not self.click_cambio_expediente():
                    logger.warning("No se pudo hacer clic en 'Cambio de Expediente'")
                
                time.sleep(0.5)
                
                # ============================================================
                # PASO 8: Iniciar bucle de expedientes
                # ============================================================
                logger.info("=" * 70)
                logger.info("PASO 8: Iniciando bucle de búsqueda de expedientes restantes")
                logger.info("=" * 70)
                
                if not self.expediente_loop_dse():
                    logger.warning("Error en expediente_loop_dse")
                
                logger.info(" PROCESO COMPLETADO EXITOSAMENTE")
                return True
            
            # ============================================================
            # MONTO ACEPTADO - Detectar avisos opcionales
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 4: MONTO ACEPTADO - Detectando avisos opcionales")
            logger.info("=" * 70)
            
            # Aviso OPCIONAL 1: "El Expediente XXX correspondiente al RUC XXX tiene X Embargos activos..."
            logger.info("Detectando aviso de embargos activos (OPCIONAL)...")
            embargo_detectado, embargo_msg = self.detect_expediente_aviso(timeout=2)
            
            if embargo_detectado:
                logger.warning(f"Aviso de embargos detectado: {embargo_msg}")
                logger.info("Presionando ALT+S...")
                pyautogui.hotkey('alt', 's')
                time.sleep(1)
                logger.info(" ALT+S presionado")
            else:
                logger.info("No se detectó aviso de embargos (continuando...)")
            
            # Aviso OPCIONAL 2: "¿ Desea Continuar ?" (con espacios específicos)
            logger.info("Detectando aviso '¿ Desea Continuar ?' (OPCIONAL)...")
            desea_continuar_detectado, desea_continuar_msg = self.detect_desea_continuar_aviso(timeout=2)
            
            if desea_continuar_detectado:
                logger.warning(f"Aviso '¿ Desea Continuar ?' detectado: {desea_continuar_msg}")
                logger.info("Presionando ALT+S...")
                pyautogui.hotkey('alt', 's')
                time.sleep(1)
                logger.info(" ALT+S presionado")
            else:
                logger.info("No se detectó '¿ Desea Continuar ?' (continuando...)")
            
            # ============================================================
            # PASO 5: Detectar "¿Desea Ud. grabar la Resolución?" (OPCIONAL)
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 5: Detectando '¿Desea Ud. grabar la Resolución Coactiva?' (OPCIONAL)")
            logger.info("=" * 70)
            
            grabar_resolucion_detectado, grabar_resolucion_msg = self.detect_grabar_resolucion_aviso(timeout=2)
            
            if grabar_resolucion_detectado:
                logger.warning(f"Aviso '¿Desea Ud. grabar Resolución?' detectado: {grabar_resolucion_msg}")
                logger.info("Presionando ALT+S...")
                pyautogui.hotkey('alt', 's')
                time.sleep(1)
                logger.info(" ALT+S presionado")
            else:
                logger.info("No se detectó '¿Desea Ud. grabar Resolución?' (continuando...)")
            
            # ============================================================
            # PASO 6: Detectar y extraer Resolución Coactiva (OBLIGATORIO)
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 6: Detectando mensaje de Resolución Coactiva (OBLIGATORIO)")
            logger.info("=" * 70)
            
            rc_detectado, rc_msg = self.detect_resolucion_coactiva_aviso(timeout=2)
            
            if rc_detectado:
                logger.warning(f"RC detectado: {rc_msg}")
                
                # Extraer número de RC
                rc_number = self.extract_resolucion_coactiva_number(rc_msg)
                
                if rc_number:
                    logger.info(f"RC extraído: {rc_number}")
                    logger.info("Guardando RC en Excel...")
                    self.update_excel_result_for_row(row_idx, rc_number)
                    logger.info(" RC guardado en Excel")
                else:
                    logger.warning("No se pudo extraer RC del mensaje")
                    self.update_excel_result_for_row(row_idx, "RC NO EXTRAÍDO")
                
                # Presionar ENTER para cerrar el mensaje de RC
                logger.info("Presionando ENTER para cerrar mensaje de RC...")
                pyautogui.press('return')
                time.sleep(1)
                logger.info(" ENTER presionado")
            else:
                logger.warning("No se detectó mensaje de RC")
                self.update_excel_result_for_row(row_idx, "RC NO DETECTADO")
            
            logger.info("=" * 70)
            logger.info(" FLUJO DE MONTO ACEPTADO - INICIANDO BUCLE PARA SIGUIENTE EXPEDIENTE")
            logger.info("=" * 70)
            
            # ============================================================
            # PASO 7: Presionar ALT+C para regresar al menú
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 7: Presionando ALT+C para regresar al menú")
            logger.info("=" * 70)
            
            logger.info("Presionando ALT+C para regresar al menú...")
            time.sleep(0.5)
            pyautogui.hotkey('alt', 'c')
            time.sleep(1)
            logger.info(" ALT+C presionado correctamente")
            
            # ============================================================
            # PASO 8: Eliminar desplazamientos del menú para acceder a "Accesos"
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 8: Eliminando desplazamientos del menú")
            logger.info("=" * 70)
            
            # Clic en "Trabar Embargo" para eliminar un desplazamiento
            logger.info("Haciendo clic en 'Trabar Embargo' para eliminar desplazamiento...")
            if not self.click_trabar_embargo():
                logger.warning("No se pudo hacer clic en 'Trabar Embargo'")
            time.sleep(0.5)
            
            # Clic en "Proceso de Embargo" para eliminar otro desplazamiento
            logger.info("Haciendo clic en 'Proceso de Embargo' para eliminar desplazamiento...")
            if not self.click_proceso_embargo():
                logger.warning("No se pudo hacer clic en 'Proceso de Embargo'")
            time.sleep(0.5)
            
            logger.info(" Desplazamientos eliminados - 'Accesos' ahora visible en el mismo nivel")
            
            # ============================================================
            # PASO 9: Hacer clic en "Accesos"
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 9: Haciendo clic en 'Accesos'")
            logger.info("=" * 70)
            
            # Buscar y hacer clic en "Accesos" directamente (sin desplazamiento)
            logger.info("Buscando 'Accesos' en el menú...")
            if not self._click_accesos_direct():
                logger.warning("No se pudo hacer clic en 'Accesos'")
            
            time.sleep(0.5)
            
            # ============================================================
            # PASO 10: Hacer doble clic en "Cambio de Expediente"
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 10: Haciendo doble clic en 'Cambio de Expediente'")
            logger.info("=" * 70)
            
            if not self.click_cambio_expediente():
                logger.warning("No se pudo hacer clic en 'Cambio de Expediente'")
            
            time.sleep(0.5)
            
            # ============================================================
            # PASO 11: Iniciar bucle de expedientes
            # ============================================================
            logger.info("=" * 70)
            logger.info("PASO 11: Iniciando bucle de búsqueda de expedientes restantes")
            logger.info("=" * 70)
            
            if not self.expediente_loop_dse():
                logger.warning("Error en expediente_loop_dse")
            
            logger.info(" PROCESO COMPLETADO EXITOSAMENTE")
            return True
        
        except Exception as e:
            logger.error(f"Error en fill_monto_loop: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return False
    
    def expediente_loop_dse(self):
        """
        Bucle de búsqueda de expedientes FLEXIBLE que detecta el tipo de medida en cada expediente.
        
        IMPORTANTE: Después de procesar el primer expediente (DSE o IEI),
        el campo "Cambio de Expediente" está listo para ingresar el siguiente expediente.
        Este bucle continúa desde ese punto.
        
        CARACTERÍSTICA CLAVE: Detecta dinámicamente el TIPO DE MEDIDA de CADA expediente
        y ejecuta el flujo correspondiente (DSE o IEI).
        
        Flujo para cada expediente:
        1. Lee el siguiente expediente del Excel
        2. Ingresa el expediente
        3. Presiona ENTER para validar
        4. Verifica si es válido (sin error)
        5. Si es válido: 
           - Presiona ALT+A para validar ejecutor
           - Lee el TIPO DE MEDIDA de ese expediente
           - Si es DSE: Hace clic en "Trabar Depósito sin Extracción"fill_monto_loop()
           - Si es IEI: Hace clic en "Trabar Intervención en Información"fill_interventor_and_plazo_loop()
        6. Si es inválido: Marca como "EXP. INVALIDO" y continúa
        7. Repite hasta procesar todos los expedientes
        
        NOTA: El tipo de medida puede cambiar entre expedientes (ej: DSEIEIDSE)
        """
        logger.info("\n" + "=" * 70)
        logger.info("INICIANDO BUCLE DE EXPEDIENTES FLEXIBLE (DETECTANDO TIPO POR EXPEDIENTE)")
        logger.info("=" * 70)
        
        try:
            excel_file = SCRIPT_DIR / "EXPEDIENTES.xlsx"
            expedientes = pd.read_excel(excel_file, engine="openpyxl", dtype=str)
            
            # Empezar desde el segundo expediente (índice 1)
            for idx in range(self.primer_expediente_idx + 1, len(expedientes)):
                exp_actual = str(expedientes.iloc[idx]["EXPEDIENTE"]).strip()
                
                logger.info(f"\n{'=' * 70}")
                logger.info(f"PROCESANDO EXPEDIENTE {idx + 1} DE {len(expedientes)}: {exp_actual}")
                logger.info(f"{'=' * 70}")
                
                # ============================================================
                # PASO 1: Digitar el expediente
                # ============================================================
                logger.info("PASO 1: Digitando expediente...")
                logger.info(f"Expediente: {exp_actual}")
                pyautogui.write(exp_actual, interval=0.05)
                time.sleep(0.5)
                
                # ============================================================
                # PASO 2: Presionar ENTER para validar el expediente
                # ============================================================
                logger.info("PASO 2: Presionando ENTER para validar expediente...")
                pyautogui.press('return')
                time.sleep(1)
                logger.info(" ENTER presionado")
                
                # ============================================================
                # PASO 3: Verificar si el expediente es válido
                # ============================================================
                logger.info("PASO 3: Verificando si el expediente es válido...")
                
                # Detectar error de expediente
                error_detectado, error_msg = self.detect_expediente_error(timeout=2)
                
                if error_detectado:
                    logger.error(f"Expediente inválido: {error_msg}")
                    
                    # Marcar como inválido en Excel
                    self.mark_invalid_expediente_in_results(idx)
                    logger.warning(f"Expediente marcado como EXP. INVALIDO en fila {idx + 2}")
                    
                    # Presionar ENTER para cerrar diálogo de error
                    logger.info("Presionando ENTER para cerrar diálogo de error...")
                    pyautogui.press('return')
                    time.sleep(1)
                    
                    # Borrar el expediente inválido (Ctrl+A, Delete)
                    logger.info("Borrando expediente inválido...")
                    pyautogui.hotkey('ctrl', 'a')
                    time.sleep(0.2)
                    pyautogui.press('delete')
                    time.sleep(0.5)
                    
                    # Continuar con el siguiente expediente
                    logger.info("Continuando con el siguiente expediente...")
                    continue
                else:
                    logger.info(" Expediente es VÁLIDO")
                
                # ============================================================
                # PASO 4: Presionar ALT+A para validar ejecutor
                # ============================================================
                logger.info("PASO 4: Presionando ALT+A para validar ejecutor...")
                pyautogui.hotkey('alt', 'a')
                time.sleep(1)
                logger.info(" ALT+A presionado correctamente")
                
                # ============================================================
                # PASO 5: VALIDAR QUE EL EXPEDIENTE TENGA TODOS LOS DATOS
                # ============================================================
                logger.info("=" * 70)
                logger.info("PASO 5: VALIDANDO DATOS DEL EXPEDIENTE")
                logger.info("=" * 70)
                
                es_valido, mensaje_error = self.validate_expediente_row(expedientes, idx)
                
                if not es_valido:
                    logger.error(f"Datos incompletos: {mensaje_error}")
                    
                    # Marcar con el motivo específico en Excel
                    self.mark_invalid_expediente_in_results(idx, mensaje_error)
                    
                    # Limpiar campo y continuar
                    logger.info("Limpiando campo de expediente...")
                    pyautogui.hotkey('ctrl', 'backspace')
                    time.sleep(0.2)
                    
                    logger.info("Continuando con el siguiente expediente...")
                    continue
                
                logger.info(" Todos los datos del expediente son válidos")
                
                # ============================================================
                logger.info("=" * 70)
                logger.info("PASO 6: DETECTANDO TIPO DE MEDIDA PARA ESTE EXPEDIENTE")
                logger.info("=" * 70)
                
                # Leer el TIPO DE MEDIDA de la fila actual
                tipo_medida_actual = str(expedientes.iloc[idx]["TIPO DE MEDIDA"]).strip().upper()
                logger.info(f"Tipo de Medida detectado para expediente {idx + 1}: {tipo_medida_actual}")
                
                # Determinar si es IEI o DSE
                if "IEI" in tipo_medida_actual:
                    medida_tipo = "IEI"
                    logger.info(" Es IEI (Trabar Intervención en Información)")
                elif "DSE" in tipo_medida_actual:
                    medida_tipo = "DSE"
                    logger.info(" Es DSE (Trabar Depósito sin Extracción)")
                else:
                    logger.error(f"Tipo de medida no reconocido: {tipo_medida_actual}")
                    self.mark_invalid_expediente_in_results(idx, f"TIPO NO VÁLIDO: {tipo_medida_actual}")
                    continue
                
                # ============================================================
                # PASO 7: Hacer clic en "Trabar Embargo"
                # ============================================================
                logger.info("PASO 7: Haciendo clic en 'Trabar Embargo'...")
                if not self.click_trabar_embargo():
                    logger.warning("No se pudo hacer clic en 'Trabar Embargo'")
                time.sleep(0.5)
                
                # ============================================================
                # PASO 8: EJECUTAR FLUJO SEGÚN TIPO DE MEDIDA
                # ============================================================
                logger.info("=" * 70)
                logger.info(f"PASO 8: EJECUTANDO FLUJO DE {medida_tipo}")
                logger.info("=" * 70)
                
                if medida_tipo == "DSE":
                    # ============================================================
                    # FLUJO DSE: Trabar Depósito sin Extracciónfill_monto_loop()
                    # ============================================================
                    logger.info("Ejecutando flujo DSE...")
                    
                    # Hacer doble clic en "Trabar Depósito sin Extracción"
                    logger.info("Haciendo doble clic en 'Trabar Depósito sin Extracción'...")
                    if not self.click_trabar_deposito_sin_extraccion():
                        logger.warning("No se pudo hacer clic en 'Trabar Depósito sin Extracción'")
                    
                    # Manejar el aviso que puede aparecer
                    logger.info("Manejando posible aviso...")
                    if not self.handle_trabar_deposito_aviso():
                        logger.warning("No se pudo manejar el aviso")
                    
                    time.sleep(0.5)
                    
                    # Rellenar MONTO con fill_monto_loop()
                    logger.info("Rellenando campo MONTO...")
                    if not self.fill_monto_loop(idx):
                        logger.warning(f"Error rellenando MONTO para expediente en fila {idx + 2}")
                
                elif medida_tipo == "IEI":
                    # ============================================================
                    # FLUJO IEI: Trabar Intervención en Informaciónfill_interventor_and_plazo_loop()
                    # ============================================================
                    logger.info("Ejecutando flujo IEI...")
                    
                    # Hacer doble clic en "Trabar Intervención en Información"
                    logger.info("Haciendo doble clic en 'Trabar Intervención en Información'...")
                    if not self.click_trabar_intervencion_informacion():
                        logger.warning("No se pudo hacer clic en 'Trabar Intervención en Información'")
                    
                    # Manejar el aviso que puede aparecer
                    logger.info("Manejando posible aviso...")
                    if not self.handle_trabar_intervencion_aviso():
                        logger.warning("No se pudo manejar el aviso")
                    
                    time.sleep(0.5)
                    
                    # Rellenar INTERVENTOR y PLAZO con fill_interventor_and_plazo_loop()
                    logger.info("Rellenando campos INTERVENTOR y PLAZO...")
                    if not self.fill_interventor_and_plazo_loop(idx):
                        logger.warning(f"Error rellenando INTERVENTOR/PLAZO para expediente en fila {idx + 2}")
                
                logger.info(f" Expediente {idx + 1} procesado exitosamente")
            
            logger.info("\n" + "=" * 70)
            logger.info("BUCLE DE EXPEDIENTES COMPLETADO EXITOSAMENTE")
            logger.info("=" * 70)
            return True
        
        except Exception as e:
            logger.error(f"Error en expediente_loop_dse: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return False
    
    def get_expedientes_grouped_by_dependencia(self):
        """
        Lee el Excel y agrupa expedientes por dependencia.

        Retorna:
            - dict con estructura: {
                "21": [filas de expedientes con dependencia 21],
                "23": [filas de expedientes con dependencia 23]
              }
            - Orden de las dependencias: la DEPENDENCIA más usada (mayor cantidad de expedientes) inicia.
              En caso de empate, se usa la dependencia que apareció primero en el Excel como desempate.

        Ejemplo:
            Si hay 3 filas con dependencia 23 y 2 filas con dependencia 21,
            el orden devuelto será ["23", "21"] (23 inicia porque tiene más expedientes).
        """
        try:
            excel_file = SCRIPT_DIR / "EXPEDIENTES.xlsx"
            expedientes = pd.read_excel(excel_file, engine="openpyxl", dtype=str)
            
            grupos = {}
            # registro del primer indice donde aparece cada dependencia (para desempates)
            primera_aparicion = {}

            for idx, fila in expedientes.iterrows():
                dependencia = str(fila.get("DEPENDENCIA", "")).strip()

                # Extraer "21" o "23" de la dependencia
                dep_key = None
                if "21" in dependencia:
                    dep_key = "21"
                elif "23" in dependencia:
                    dep_key = "23"
                else:
                    # ignorar filas sin dependencia valida
                    continue

                if dep_key not in grupos:
                    grupos[dep_key] = []
                    primera_aparicion[dep_key] = idx

                grupos[dep_key].append((idx, fila))

            # Ordenar dependencias por cantidad descendente; en empate, por primera aparición
            orden_dependencias = sorted(
                list(grupos.keys()),
                key=lambda k: (-len(grupos[k]), primera_aparicion.get(k, float('inf')))
            )
            
            logger.info(f"\n Expedientes agrupados por dependencia:")
            for dep_key in orden_dependencias:
                logger.info(f"  • Dependencia {dep_key}: {len(grupos[dep_key])} expedientes")
            
            logger.info(f" Orden de procesamiento: {' → '.join(orden_dependencias)}")
            
            return grupos, orden_dependencias
        
        except Exception as e:
            logger.error(f"Error agrupando expedientes: {str(e)}")
            return {}, []
    
    def process_dependencia_batch(self, dependencia, expedientes_grupo, is_first=False, is_last=False):
        """
        Procesa un lote de expedientes de una dependencia.
        
        Flujo:
        1. Si es PRIMER lote: Abre app + Login
        2. Para CADA expediente:
           a. Valida si el expediente tiene datos completos
           b. Si es inválido: Marca en Excel y continúa al siguiente
           c. Si es válido:
              - PRIMER expediente del lote: Ejecuta click_cobranza_coactiva() (desde cero)
              - SIGUIENTES expedientes del lote: Ejecuta click_cambio_expediente() (reusa flujo)
        3. ALT+F4 para cerrar app (SIEMPRE)
        
        IMPORTANTE: 
        - ALT+F4 SIEMPRE cierra la app completamente
        - Esto permite loguear nuevamente con la siguiente dependencia
        - "Desde cero" SOLO aplica al PRIMER expediente de cada lote (después de ALT+F4)
        
        Args:
            dependencia: "21" o "23"
            expedientes_grupo: Lista de tuplas (idx, fila) del Excel
            is_first: True si es el primer lote (abre app + login)
            is_last: True si es el último lote
        """
        try:
            logger.info("\n" + "=" * 70)
            logger.info(f"PROCESANDO LOTE DE DEPENDENCIA {dependencia}")
            logger.info(f"Expedientes a procesar: {len(expedientes_grupo)}")
            logger.info(f"Primer lote: {is_first} | Último lote: {is_last}")
            logger.info("=" * 70)
            
            # Cargar datos del Excel para validación
            excel_file = SCRIPT_DIR / "EXPEDIENTES.xlsx"
            expedientes_df = pd.read_excel(excel_file, engine="openpyxl", dtype=str)
            
            # PASO 1: Si es el primer lote, abrir app y hacer login
            if is_first:
                logger.info("\n[PRIMER LOTE] Abriendo aplicación...")
                if not self.open_application():
                    logger.error("No se pudo abrir la aplicación")
                    return False
                
                logger.info("Esperando carga de la aplicación...")
                time.sleep(3)
                
                logger.info("Procediendo con el login...")
                if not self.login():
                    logger.warning("No se pudo completar el login")
                    return False
            else:
                # Para lotes posteriores, app será abierta nuevamente después de ALT+F4
                logger.info("\n[LOTE POSTERIOR] La app fue cerrada, reabriendo...")
                if not self.open_application():
                    logger.error("No se pudo abrir la aplicación")
                    return False
                
                logger.info("Esperando carga de la aplicación...")
                time.sleep(3)
                
                logger.info("Procediendo con el login...")
                if not self.login():
                    logger.warning("No se pudo completar el login")
                    return False
            
            # PASO 2: Procesar cada expediente del grupo
            logger.info(f"\nProcesando {len(expedientes_grupo)} expedientes de dependencia {dependencia}...")
            
            for idx, (row_idx, fila) in enumerate(expedientes_grupo, 1):
                logger.info(f"\n" + "=" * 70)
                logger.info(f"EXPEDIENTE [{idx}/{len(expedientes_grupo)}] - Fila {row_idx + 2}")
                logger.info("=" * 70)
                
                # ============================================================
                # VALIDAR EXPEDIENTE PRIMERO
                # ============================================================
                es_valido, error_msg = self.validate_expediente_row(expedientes_df, row_idx)
                
                if not es_valido:
                    # Marcar como inválido en Excel y continuar
                    logger.warning(f"✗ Expediente inválido: {error_msg}")
                    self.mark_invalid_expediente_in_results(row_idx, error_msg)
                    continue
                
                # ============================================================
                # EXPEDIENTE VÁLIDO: Procesar según posición en lote
                # ============================================================
                if idx == 1:
                    # PRIMER expediente del lote: Cobranza Coactiva (DESDE CERO)
                    logger.info("→ PRIMER expediente del lote: Ejecutando desde Cobranza Coactiva...")
                    resultado = self.click_cobranza_coactiva()
                else:
                    # SIGUIENTES expedientes: decidir modo de ingreso según
                    # si el expediente anterior quedó completamente procesado.
                    logger.info("→ Expediente siguiente en mismo lote: Decidiendo modo de ingreso...")
                    # Si el expediente anterior finalizó correctamente (RC o MONTO MAYOR),
                    # podemos usar el flujo de cambio de expediente en la UI.
                    if getattr(self, 'last_exp_completed', False):
                        logger.info("Expediente anterior completado → usando click_cambio_expediente()")
                        resultado = self.click_cambio_expediente()
                        if resultado:
                            # Tras el cambio, ingresar y validar el expediente concreto
                            enter_ok = self.enter_specific_expediente(row_idx)
                            if enter_ok:
                                resultado = self.validate_executor()
                                if resultado:
                                    resultado = self.handle_post_embargo_flow()
                            else:
                                resultado = False
                        else:
                            logger.warning("No se pudo ejecutar click_cambio_expediente()")
                            # Intentar ingresar directamente sin cambio como fallback
                            enter_ok = self.enter_specific_expediente(row_idx)
                            if enter_ok:
                                resultado = self.validate_executor()
                                if resultado:
                                    resultado = self.handle_post_embargo_flow()
                            else:
                                resultado = False
                    else:
                        # Si el anterior NO completó su flujo (p. ej. fue inválido),
                        # NO debemos hacer click de cambio de expediente en la UI.
                        logger.info("Expediente anterior NO completado → ingresando directamente sin cambio de UI")
                        enter_ok = self.enter_specific_expediente(row_idx)
                        if enter_ok:
                            resultado = self.validate_executor()
                            if resultado:
                                resultado = self.handle_post_embargo_flow()
                        else:
                            resultado = False
                
                if resultado:
                    logger.info(f"✓ [{idx}/{len(expedientes_grupo)}] Expediente procesado")
                else:
                    logger.warning(f"✗ [{idx}/{len(expedientes_grupo)}] Error procesando expediente")
            
            logger.info(f"\n✓ Lote de dependencia {dependencia} completado")
            
            # PASO 3: Cerrar app con ALT+F4 (SIEMPRE, incluso en último lote)
            logger.info("\nCerrando aplicación con ALT+F4...")
            try:
                pyautogui.hotkey('alt', 'f4')
                time.sleep(2)
                logger.info(" ✓ Aplicación cerrada")
            except Exception as e:
                logger.warning(f" Error en ALT+F4: {e}")
            
            return True
        
        except Exception as e:
            logger.error(f"Error en process_dependencia_batch: {str(e)}")
            return False
    
    def run(self):
        """
        Ejecuta la automatización procesando múltiples dependencias con UNA SOLA APP abierta.

        Flujo:
        1. Agrupa expedientes por dependencia (21 y 23) y ordena para iniciar por la DEPENDENCIA más usada
           (mayor cantidad de expedientes). En empate, inicia con la que aparece primero en el Excel.
        2. Abre la aplicación UNA SOLA VEZ
        3. Para cada dependencia en el orden calculado:
           a. Hace login con esa dependencia (primera vez abre + login, siguientes solo login)
           b. Procesa todos sus expedientes
           c. Cierra sesión (ALT+F4) si hay más dependencias después
        4. Al final cierra completamente la aplicación
        """
        try:
            logger.info("\n" + "=" * 70)
            logger.info("INICIALIZANDO PROCESAMIENTO MULTI-DEPENDENCIA (UNA APP)")
            logger.info("=" * 70)
            
            # PASO 1: Agrupar expedientes por dependencia
            grupos_expedientes, orden_deps = self.get_expedientes_grouped_by_dependencia()
            
            if not orden_deps:
                logger.error("No se encontraron expedientes válidos")
                return False
            
            # PASO 2: Procesar cada dependencia en orden
            total_grupos = len(orden_deps)
            
            for grupo_idx, dependencia in enumerate(orden_deps, 1):
                logger.info("\n" + "=" * 70)
                logger.info(f"GRUPO {grupo_idx}/{total_grupos}: DEPENDENCIA {dependencia}")
                logger.info("=" * 70)
                
                # Determinar si es el primer o último lote
                is_first = (grupo_idx == 1)
                is_last = (grupo_idx == total_grupos)
                
                # Procesar todos los expedientes de esta dependencia
                expedientes_grupo = grupos_expedientes[dependencia]
                
                if not self.process_dependencia_batch(
                    dependencia, 
                    expedientes_grupo, 
                    is_first=is_first, 
                    is_last=is_last
                ):
                    logger.error(f"Error procesando dependencia {dependencia}")
                    # Continuar con la siguiente dependencia de todas formas
                    continue
            
            logger.info("\n" + "=" * 70)
            logger.info("PROCESAMIENTO COMPLETADO")
            logger.info(f"Se procesaron {total_grupos} dependencia(s)")
            logger.info("=" * 70)
            return True
        
        except Exception as e:
            logger.error(f"Error en run (multi-dependencia): {str(e)}")
            return False

def main():
    """Función principal"""
    automation = RSIRATAutomation32()
    result = automation.run()
    
    if result:
        logger.info("\nEl proceso se completó sin errores.")
    else:
        logger.error("\nEl proceso finalizó con errores. Revisa el log para más detalles.")
    
    return result


if __name__ == "__main__":
    main()
